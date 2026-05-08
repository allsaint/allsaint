# -------------------- IMPORTS --------------------
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
import os
from datetime import datetime, date, timedelta
from calendar import month_name
import uuid
import json
import io
import bcrypt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import logging
import sys


# -------------------- FLASK APP SETUP --------------------
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "super_secret_key_change_later")

# Configure logging
if not app.debug:
    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setLevel(logging.INFO)
    app.logger.addHandler(stream_handler)
    
app.logger.setLevel(logging.INFO)

# -------------------- DATABASE CONFIGURATION --------------------
# Use SQLite database
DATABASE_PATH = os.environ.get('DATABASE_PATH', 'hospital.db')

def get_db_connection():
    """Establish SQLite database connection."""
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row  # Access columns by name
        # Enable foreign keys
        conn.execute("PRAGMA foreign_keys = ON")
        return conn
    except Exception as e:
        app.logger.error(f"Database connection error: {e}")
        return None

# Session configuration
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    PERMANENT_SESSION_LIFETIME=timedelta(days=7)
)

# Health check endpoint
@app.route('/health')
def health_check():
    """Health check endpoint."""
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("SELECT 1")
            cursor.close()
            conn.close()
            return jsonify({"status": "healthy", "database": "connected"}), 200
        else:
            return jsonify({"status": "unhealthy", "database": "disconnected"}), 500
    except Exception as e:
        app.logger.error(f"Health check failed: {e}")
        return jsonify({"status": "unhealthy", "error": str(e)}), 500

# -------------------- DATABASE INITIALIZATION --------------------
def create_tables():
    """Create all necessary tables if they don't exist."""
    queries = {
        "admin_users": """
            CREATE TABLE IF NOT EXISTS admin_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                full_name VARCHAR(100),
                email VARCHAR(100),
                role VARCHAR(50) DEFAULT 'Admin',
                is_super_admin BOOLEAN DEFAULT 0,
                is_active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                created_by INTEGER REFERENCES admin_users(id),
                last_login TIMESTAMP
            );
        """,
        
        "cashier_users": """
            CREATE TABLE IF NOT EXISTS cashier_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                full_name VARCHAR(100),
                email VARCHAR(100),
                is_active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                created_by INTEGER REFERENCES admin_users(id),
                last_login TIMESTAMP
            );
        """,
        
        "admin_audit_logs": """
            CREATE TABLE IF NOT EXISTS admin_audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                admin_id INTEGER REFERENCES admin_users(id),
                action VARCHAR(100) NOT NULL,
                details TEXT,
                ip_address VARCHAR(45),
                user_agent TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        
        "pharmacists": """
            CREATE TABLE IF NOT EXISTS pharmacists (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                full_name VARCHAR(100),
                is_active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                created_by INTEGER REFERENCES admin_users(id)
            );
        """,
        
        "billing_users": """
            CREATE TABLE IF NOT EXISTS billing_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                full_name VARCHAR(100),
                is_active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                created_by INTEGER REFERENCES admin_users(id)
            );
        """,
        
        "drugs": """
            CREATE TABLE IF NOT EXISTS drugs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name VARCHAR(100) NOT NULL,
                strength VARCHAR(50) NOT NULL,
                unit_price DECIMAL(10, 2) NOT NULL,
                stock_quantity INT NOT NULL,
                expiry_date DATE NOT NULL,
                low_stock_threshold INT DEFAULT 20,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        
        "drug_sales": """
            CREATE TABLE IF NOT EXISTS drug_sales (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                receipt_no VARCHAR(50) UNIQUE NOT NULL,
                patient_name VARCHAR(100),
                patient_id VARCHAR(50),
                items TEXT NOT NULL,
                subtotal DECIMAL(10, 2) NOT NULL,
                discount DECIMAL(10, 2) DEFAULT 0.00,
                tax DECIMAL(10, 2) DEFAULT 0.00,
                grand_total DECIMAL(10, 2) NOT NULL,
                pharmacist VARCHAR(50) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        
        "receipts": """
            CREATE TABLE IF NOT EXISTS receipts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                patient_name VARCHAR(100),
                patient_id VARCHAR(50),
                subtotal DECIMAL(10, 2) NOT NULL,
                discount DECIMAL(10, 2) DEFAULT 0.00,
                tax DECIMAL(10, 2) DEFAULT 0.00,
                total_amount DECIMAL(10, 2) NOT NULL,
                grand_total DECIMAL(10, 2) NOT NULL,
                pharmacist VARCHAR(50),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        
        "receipt_items": """
            CREATE TABLE IF NOT EXISTS receipt_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                receipt_id INT NOT NULL REFERENCES receipts(id) ON DELETE CASCADE,
                drug_name VARCHAR(100) NOT NULL,
                strength VARCHAR(50) NOT NULL,
                quantity INT NOT NULL,
                unit_price DECIMAL(10, 2) NOT NULL
            );
        """,
        
        "stock_movements": """
            CREATE TABLE IF NOT EXISTS stock_movements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                drug_id INT NOT NULL REFERENCES drugs(id) ON DELETE CASCADE,
                movement_type VARCHAR(20) NOT NULL,
                quantity INT NOT NULL,
                user_id INT NOT NULL,
                note TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        
        "billing_invoice": """
            CREATE TABLE IF NOT EXISTS billing_invoice (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                patient_name VARCHAR(100) NOT NULL,
                service_type VARCHAR(100) NOT NULL,
                amount DECIMAL(10, 2) NOT NULL,
                status VARCHAR(20) DEFAULT 'UNPAID',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        
        "billing_receipt": """
            CREATE TABLE IF NOT EXISTS billing_receipt (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id INT NOT NULL REFERENCES billing_invoice(id) ON DELETE CASCADE,
                amount_paid DECIMAL(10, 2) NOT NULL,
                payment_method VARCHAR(50) NOT NULL,
                received_by VARCHAR(50) NOT NULL,
                payment_date TIMESTAMP NOT NULL
            );
        """,
        
        "payments": """
            CREATE TABLE IF NOT EXISTS payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                patient_name VARCHAR(100) NOT NULL,
                service_type VARCHAR(100) NOT NULL,
                subtotal DECIMAL(10, 2) NOT NULL,
                discount DECIMAL(10, 2) DEFAULT 0.00,
                tax DECIMAL(10, 2) DEFAULT 0.00,
                grand_total DECIMAL(10, 2) NOT NULL,
                amount_paid DECIMAL(10, 2) NOT NULL,
                balance DECIMAL(10, 2) NOT NULL,
                payment_method VARCHAR(50) NOT NULL,
                status VARCHAR(20) NOT NULL,
                payment_date DATE NOT NULL,
                recorded_by INT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        
        "users": """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        
        "hr_users": """
            CREATE TABLE IF NOT EXISTS hr_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username VARCHAR(50) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                full_name VARCHAR(100) NOT NULL,
                email VARCHAR(100),
                role VARCHAR(50) DEFAULT 'HR Staff',
                is_active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        
        "departments": """
            CREATE TABLE IF NOT EXISTS departments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name VARCHAR(100) NOT NULL,
                code VARCHAR(20) UNIQUE NOT NULL,
                description TEXT,
                head_of_dept VARCHAR(100),
                status VARCHAR(20) DEFAULT 'Active',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        
        "staff": """
            CREATE TABLE IF NOT EXISTS staff (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id VARCHAR(50) UNIQUE NOT NULL,
                first_name VARCHAR(100) NOT NULL,
                last_name VARCHAR(100) NOT NULL,
                department_id INTEGER REFERENCES departments(id),
                position VARCHAR(100) NOT NULL,
                employment_type VARCHAR(50),
                email VARCHAR(100),
                phone VARCHAR(20),
                hire_date DATE NOT NULL,
                salary DECIMAL(12, 2),
                status VARCHAR(20) DEFAULT 'Active',
                emergency_contact VARCHAR(100),
                address TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        
        "attendance": """
            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id INTEGER REFERENCES staff(id) ON DELETE CASCADE,
                date DATE NOT NULL,
                check_in TIME,
                check_out TIME,
                status VARCHAR(20),
                remarks TEXT,
                recorded_by INTEGER REFERENCES hr_users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(staff_id, date)
            );
        """,
        
        "leaves": """
            CREATE TABLE IF NOT EXISTS leaves (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id INTEGER REFERENCES staff(id) ON DELETE CASCADE,
                leave_type VARCHAR(50) NOT NULL,
                start_date DATE NOT NULL,
                end_date DATE NOT NULL,
                days_requested INTEGER NOT NULL,
                reason TEXT,
                status VARCHAR(20) DEFAULT 'Pending',
                approved_by INTEGER REFERENCES hr_users(id),
                approved_at TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        
        "schedules": """
            CREATE TABLE IF NOT EXISTS schedules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id INTEGER REFERENCES staff(id) ON DELETE CASCADE,
                schedule_date DATE NOT NULL,
                shift_type VARCHAR(50),
                start_time TIME NOT NULL,
                end_time TIME NOT NULL,
                location VARCHAR(100),
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        
        "payroll": """
            CREATE TABLE IF NOT EXISTS payroll (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id INTEGER REFERENCES staff(id) ON DELETE CASCADE,
                pay_period VARCHAR(50),
                basic_salary DECIMAL(12, 2),
                allowances DECIMAL(12, 2),
                deductions DECIMAL(12, 2),
                net_salary DECIMAL(12, 2),
                status VARCHAR(20) DEFAULT 'Pending',
                payment_date DATE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        
        "documents": """
            CREATE TABLE IF NOT EXISTS documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id INTEGER REFERENCES staff(id) ON DELETE CASCADE,
                document_type VARCHAR(50),
                document_name VARCHAR(255),
                file_path VARCHAR(500),
                uploaded_by INTEGER REFERENCES hr_users(id),
                uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """,
        
        "shift_swap_requests": """
            CREATE TABLE IF NOT EXISTS shift_swap_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                schedule_id INTEGER NOT NULL REFERENCES schedules(id) ON DELETE CASCADE,
                from_staff_id INTEGER NOT NULL REFERENCES staff(id) ON DELETE CASCADE,
                to_staff_id INTEGER NOT NULL REFERENCES staff(id) ON DELETE CASCADE,
                reason TEXT,
                status VARCHAR(20) DEFAULT 'Pending',
                requested_by INTEGER REFERENCES hr_users(id),
                requested_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                approved_by INTEGER REFERENCES hr_users(id),
                approved_at TIMESTAMP,
                reviewed_by INTEGER REFERENCES hr_users(id),
                reviewed_at TIMESTAMP,
                rejection_reason TEXT
            );
        """,
        # Add this to your create_tables() function in the queries dictionary
        "cashier_remittances": """
            CREATE TABLE IF NOT EXISTS cashier_remittances (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cashier_id INTEGER NOT NULL REFERENCES cashier_users(id),
                cashier_name VARCHAR(100) NOT NULL,
                remittance_date DATE NOT NULL,
                amount_collected DECIMAL(10, 2) NOT NULL,
                amount_remitted DECIMAL(10, 2) NOT NULL,
                balance DECIMAL(10, 2) NOT NULL,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                created_by INTEGER REFERENCES admin_users(id),
                UNIQUE(cashier_id, remittance_date)
            );
        """
    }

    conn = get_db_connection()
    if not conn:
        app.logger.error("Cannot connect to database for table creation")
        return

    cursor = conn.cursor()
    
    # Create all tables
    for table, query in queries.items():
        try:
            cursor.execute(query)
            app.logger.info(f"Table '{table}' created or already exists")
        except Exception as e:
            app.logger.error(f"Error creating table {table}: {e}")
    
    # Create indexes for better performance
    indexes = [
        "CREATE INDEX IF NOT EXISTS idx_drugs_name ON drugs(name);",
        "CREATE INDEX IF NOT EXISTS idx_drugs_expiry_date ON drugs(expiry_date);",
        "CREATE INDEX IF NOT EXISTS idx_drugs_stock_quantity ON drugs(stock_quantity);",
        "CREATE INDEX IF NOT EXISTS idx_drug_sales_created_at ON drug_sales(created_at);",
        "CREATE INDEX IF NOT EXISTS idx_drug_sales_receipt_no ON drug_sales(receipt_no);",
        "CREATE INDEX IF NOT EXISTS idx_receipts_created_at ON receipts(created_at);",
        "CREATE INDEX IF NOT EXISTS idx_receipt_items_receipt_id ON receipt_items(receipt_id);",
        "CREATE INDEX IF NOT EXISTS idx_payments_payment_date ON payments(payment_date);",
        "CREATE INDEX IF NOT EXISTS idx_payments_patient_name ON payments(patient_name);",
        "CREATE INDEX IF NOT EXISTS idx_payments_status ON payments(status);",
        "CREATE INDEX IF NOT EXISTS idx_staff_department ON staff(department_id);",
        "CREATE INDEX IF NOT EXISTS idx_staff_status ON staff(status);",
        "CREATE INDEX IF NOT EXISTS idx_attendance_staff_date ON attendance(staff_id, date);",
        "CREATE INDEX IF NOT EXISTS idx_attendance_date ON attendance(date);",
        "CREATE INDEX IF NOT EXISTS idx_leaves_staff_status ON leaves(staff_id, status);",
        "CREATE INDEX IF NOT EXISTS idx_leaves_status ON leaves(status);",
        "CREATE INDEX IF NOT EXISTS idx_schedules_staff_date ON schedules(staff_id, schedule_date);",
        "CREATE INDEX IF NOT EXISTS idx_payroll_staff_period ON payroll(staff_id, pay_period);",
        "CREATE INDEX IF NOT EXISTS idx_shift_swap_requests_status ON shift_swap_requests(status);",
        "CREATE INDEX IF NOT EXISTS idx_admin_users_username ON admin_users(username);",
        "CREATE INDEX IF NOT EXISTS idx_pharmacists_username ON pharmacists(username);",
        "CREATE INDEX IF NOT EXISTS idx_billing_users_username ON billing_users(username);",
        "CREATE INDEX IF NOT EXISTS idx_cashier_users_username ON cashier_users(username);",
        "CREATE INDEX IF NOT EXISTS idx_admin_audit_logs_admin_id ON admin_audit_logs(admin_id);",
        "CREATE INDEX IF NOT EXISTS idx_admin_audit_logs_created_at ON admin_audit_logs(created_at);",
        # Add to your indexes list
        "CREATE INDEX IF NOT EXISTS idx_cashier_remittances_date ON cashier_remittances(remittance_date);",
        "CREATE INDEX IF NOT EXISTS idx_cashier_remittances_cashier ON cashier_remittances(cashier_id, remittance_date);"
    ]
    
    for index_query in indexes:
        try:
            cursor.execute(index_query)
        except Exception as e:
            app.logger.warning(f"Could not create index: {e}")
    
    conn.commit()
    cursor.close()
    conn.close()
    
    app.logger.info("All tables created successfully")

# Template context processor
@app.context_processor
def utility_processor():
    return dict(
        min=min,
        max=max,
        abs=abs,
        round=round,
        len=len,
        int=int,
        float=float,
        str=str,
        date=date,
        datetime=datetime,
        timedelta=timedelta
    )

def create_default_users():
    """Create default users for pharmacy and billing modules."""
    default_users = {
        "pharmacists": ("pharmacist1", "pharma123"),
        "billing_users": ("billing1", "billing123")
    }

    conn = get_db_connection()
    if not conn:
        return

    cursor = conn.cursor()
    for table, (username, password) in default_users.items():
        hashed_pw = generate_password_hash(password)
        try:
            cursor.execute(f"""
                INSERT OR IGNORE INTO {table} (username, password)
                VALUES (?, ?)
            """, (username, hashed_pw))
        except Exception as e:
            app.logger.error(f"Error creating default user {username}: {e}")
    conn.commit()
    cursor.close()
    conn.close()
    

# Add these imports at the top if not already present
from datetime import date, datetime, timedelta

# Add these routes to your app.py file (after your existing billing routes)

# -------------------- CASHIER REMITTANCE ROUTES --------------------
@app.route('/billing/remittance', methods=['GET', 'POST'])
def cashier_remittance():
    """Handle cashier daily remittance submission."""
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))
    
    today = date.today()
    today_str = today.strftime('%Y-%m-%d')
    cashier_id = session["billing_user_id"]
    cashier_name = session.get("billing_username", "Cashier")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Get today's total collected amount
    cur.execute("""
        SELECT COALESCE(SUM(amount_paid), 0)
        FROM payments 
        WHERE DATE(payment_date) = DATE('now', 'localtime') AND recorded_by = ?
    """, (cashier_id,))
    
    total_collected = float(cur.fetchone()[0] or 0)
    
    # Check if already remitted today
    cur.execute("""
        SELECT id, amount_collected, amount_remitted, balance, notes
        FROM cashier_remittances
        WHERE cashier_id = ? AND remittance_date = DATE('now', 'localtime')
    """, (cashier_id,))
    
    existing_remittance = cur.fetchone()
    
    if request.method == 'POST':
        amount_remitted = float(request.form.get('amount_remitted', 0))
        notes = request.form.get('notes', '')
        
        if amount_remitted <= 0:
            flash("Please enter a valid remittance amount", "danger")
            return redirect(url_for('cashier_remittance'))
        
        if amount_remitted > total_collected:
            flash(f"Remittance amount (₦{amount_remitted:,.2f}) cannot exceed total collected (₦{total_collected:,.2f})", "danger")
            return redirect(url_for('cashier_remittance'))
        
        balance = total_collected - amount_remitted
        
        try:
            if existing_remittance:
                cur.execute("""
                    UPDATE cashier_remittances
                    SET amount_collected = ?,
                        amount_remitted = ?,
                        balance = ?,
                        notes = ?
                    WHERE id = ?
                """, (total_collected, amount_remitted, balance, notes, existing_remittance[0]))
                flash("Remittance updated successfully!", "success")
            else:
                cur.execute("""
                    INSERT INTO cashier_remittances 
                    (cashier_id, cashier_name, remittance_date, amount_collected, 
                     amount_remitted, balance, notes)
                    VALUES (?, ?, DATE('now', 'localtime'), ?, ?, ?, ?)
                """, (cashier_id, cashier_name,
                      total_collected, amount_remitted, balance, notes))
                flash("Remittance submitted successfully!", "success")
            
            conn.commit()
            return redirect(url_for('cashier_remittance_history'))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error saving remittance: {e}")
            flash(f"Error saving remittance: {str(e)}", "danger")
        
        finally:
            cur.close()
            conn.close()
            return redirect(url_for('cashier_remittance'))
    
    # GET request — build current_remittance for template
    current_remittance = None
    if existing_remittance:
        current_remittance = {
            'id': existing_remittance[0],
            'amount_collected': float(existing_remittance[1]),
            'amount_remitted': float(existing_remittance[2]),
            'balance': float(existing_remittance[3]),
            'notes': existing_remittance[4]
        }
    
    cur.close()
    conn.close()
    
    return render_template(
        "cashier_remittance.html",
        today=today,
        total_collected=total_collected,
        current_remittance=current_remittance,
        cashier_name=cashier_name,
        hospital_name="All Saint Medical Center Nsukka, Enugu State"
    )
    
        
    # For GET request
    current_remittance = None
    if existing_remittance:
        current_remittance = {
            'id': existing_remittance[0],
            'amount_collected': float(existing_remittance[1]),
            'amount_remitted': float(existing_remittance[2]),
            'balance': float(existing_remittance[3]),
            'notes': existing_remittance[4]
        }
    
    cur.close()
    conn.close()
    
    return render_template(
        "cashier_remittance.html",
        today=today,
        total_collected=total_collected,
        current_remittance=current_remittance,
        cashier_name=cashier_name,
        hospital_name="All Saint Medical Center Nsukka, Enugu State"
    )


@app.route('/billing/remittance/history')
def cashier_remittance_history():
    """View remittance history for the current cashier."""
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))
    
    cashier_id = session["billing_user_id"]
    cashier_name = session.get("billing_username", "Cashier")
    
    page = request.args.get("page", 1, type=int)
    per_page = 20
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get total count
        cur.execute("""
            SELECT COUNT(*) FROM cashier_remittances
            WHERE cashier_id = ?
        """, (cashier_id,))
        total_items = cur.fetchone()[0]
        
        # Get paginated history
        offset = (page - 1) * per_page
        cur.execute("""
            SELECT id, remittance_date, amount_collected, amount_remitted, 
                   balance, notes, created_at
            FROM cashier_remittances
            WHERE cashier_id = ?
            ORDER BY remittance_date DESC
            LIMIT ? OFFSET ?
        """, (cashier_id, per_page, offset))
        
        remittances = cur.fetchall()
        
        # Calculate summary statistics
        cur.execute("""
            SELECT 
                COUNT(*) as total_remittances,
                COALESCE(SUM(amount_collected), 0) as total_collected,
                COALESCE(SUM(amount_remitted), 0) as total_remitted,
                COALESCE(SUM(balance), 0) as total_balance
            FROM cashier_remittances
            WHERE cashier_id = ?
        """, (cashier_id,))
        
        summary = cur.fetchone()
        
    except Exception as e:
        app.logger.error(f"Error fetching remittance history: {e}")
        remittances = []
        total_items = 0
        summary = (0, 0, 0, 0)
        flash("Error loading remittance history", "danger")
    
    finally:
        cur.close()
        conn.close()
    
    total_pages = (total_items + per_page - 1) // per_page if total_items > 0 else 1
    
    return render_template(
        "cashier_remittance_history.html",
        remittances=remittances,
        summary={
            'total_remittances': summary[0],
            'total_collected': float(summary[1]),
            'total_remitted': float(summary[2]),
            'total_balance': float(summary[3])
        },
        page=page,
        total_pages=total_pages,
        cashier_name=cashier_name,
        hospital_name="All Saint Medical Center Nsukka, Enugu State"
    )


@app.route('/billing/daily-report')
def billing_daily_report():
    """Show daily report with remittance information."""
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))
    
    selected_date = request.args.get("date", date.today().strftime('%Y-%m-%d'))
    cashier_id = session["billing_user_id"]
    cashier_name = session.get("billing_username", "Cashier")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get today's collections
        cur.execute("""
            SELECT 
                COUNT(*) as transaction_count,
                COALESCE(SUM(amount_paid), 0) as total_collected,
                COALESCE(AVG(amount_paid), 0) as avg_transaction,
                COUNT(DISTINCT payment_method) as payment_methods_used
            FROM payments 
            WHERE DATE(payment_date) = ? AND recorded_by = ?
        """, (selected_date, cashier_id))
        
        collections = cur.fetchone()
        
        # Get payment method breakdown
        cur.execute("""
            SELECT 
                payment_method,
                COUNT(*) as count,
                COALESCE(SUM(amount_paid), 0) as total
            FROM payments 
            WHERE DATE(payment_date) = ? AND recorded_by = ?
            GROUP BY payment_method
            ORDER BY total DESC
        """, (selected_date, cashier_id))
        
        payment_breakdown = cur.fetchall()
        
        # Get remittance for the day
        cur.execute("""
            SELECT amount_collected, amount_remitted, balance, notes, created_at
            FROM cashier_remittances
            WHERE cashier_id = ? AND remittance_date = ?
        """, (cashier_id, selected_date))
        
        remittance = cur.fetchone()
        
        # Get detailed transactions for the day
        cur.execute("""
            SELECT id, patient_name, service_type, amount_paid, 
                   payment_method, status, payment_date, created_at
            FROM payments 
            WHERE DATE(payment_date) = ? AND recorded_by = ?
            ORDER BY created_at DESC
            LIMIT 50
        """, (selected_date, cashier_id))
        
        transactions = cur.fetchall()
        
        # Check if remittance is pending
        remittance_status = "pending"
        remittance_data = None
        
        if remittance:
            remittance_status = "completed"
            remittance_data = {
                'amount_collected': float(remittance[0]),
                'amount_remitted': float(remittance[1]),
                'balance': float(remittance[2]),
                'notes': remittance[3],
                'created_at': remittance[4]
            }
            
            # Check if balance matches current collections
            if remittance_data['amount_collected'] != float(collections[1]):
                remittance_status = "mismatch"
        
    except Exception as e:
        app.logger.error(f"Error generating daily report: {e}")
        collections = (0, 0, 0, 0)
        payment_breakdown = []
        remittance = None
        transactions = []
        remittance_status = "pending"
        remittance_data = None
        flash("Error generating daily report", "danger")
    
    finally:
        cur.close()
        conn.close()
    
    # Parse the selected date
    try:
        parsed_date = datetime.strptime(selected_date, '%Y-%m-%d')
    except:
        parsed_date = datetime.now()
    
    return render_template(
        "billing_daily_report.html",
        selected_date=parsed_date,
        collections={
            'transaction_count': collections[0],
            'total_collected': float(collections[1]),
            'avg_transaction': float(collections[2]),
            'payment_methods_used': collections[3]
        },
        payment_breakdown=payment_breakdown,
        remittance_status=remittance_status,
        remittance=remittance_data,
        transactions=transactions,
        cashier_name=cashier_name,
        hospital_name="All Saint Medical Center Nsukka, Enugu State"
    )


@app.route('/billing/api/today-collection')
def api_today_collection():
    """API endpoint to get today's collection and remittance status."""
    if "billing_user_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    cashier_id = session["billing_user_id"]
    today = date.today()
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get today's total collection
        cur.execute("""
            SELECT COALESCE(SUM(amount_paid), 0)
            FROM payments 
            WHERE DATE(payment_date) = DATE('now', 'localtime') AND recorded_by = ?
        """, (cashier_id,))
        
        total_collected = float(cur.fetchone()[0] or 0)
        
        # Get today's transaction count
        cur.execute("""
            SELECT COUNT(*)
            FROM payments 
            WHERE DATE(payment_date) = DATE('now', 'localtime') AND recorded_by = ?
        """, (cashier_id,))
        
        transaction_count = cur.fetchone()[0] or 0
        
        # Get today's remittance if exists
        cur.execute("""
            SELECT amount_remitted, balance, notes
            FROM cashier_remittances
            WHERE cashier_id = ? AND remittance_date = DATE('now', 'localtime')
        """, (cashier_id,))
        
        remittance_row = cur.fetchone()
        remittance = None
        if remittance_row:
            remittance = {
                'amount_remitted': float(remittance_row[0]),
                'balance': float(remittance_row[1]),
                'notes': remittance_row[2]
            }
        
        return jsonify({
            "success": True,
            "total_collected": total_collected,
            "transaction_count": transaction_count,
            "remittance": remittance
        })
        
    except Exception as e:
        app.logger.error(f"Error fetching today's collection: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    
    finally:
        cur.close()
        conn.close()
        
                    
def create_nkiru_user():
    """Create Nkiru as a default cashier/billing user."""
    conn = get_db_connection()
    if not conn:
        return
    
    cursor = conn.cursor()
    
    try:
        # Check if user already exists in cashier_users
        cursor.execute("SELECT id FROM cashier_users WHERE username = ?", ("Cashier1",))
        if not cursor.fetchone():
            hashed_pw = generate_password_hash("Nkiru1@allsaints")
            cursor.execute("""
                INSERT INTO cashier_users (username, password, full_name, is_active)
                VALUES (?, ?, ?, ?)
            """, ("Cashier1", hashed_pw, "Nkiru", 1))
            app.logger.info("Cashier user 'Nkiru' created successfully")
        else:
            app.logger.info("Cashier user 'Nkiru' already exists")
        
        # Check if user already exists in billing_users
        cursor.execute("SELECT id FROM billing_users WHERE username = ?", ("Cashier1",))
        if not cursor.fetchone():
            hashed_pw = generate_password_hash("Nkiru1@allsaints")
            cursor.execute("""
                INSERT INTO billing_users (username, password, full_name, is_active)
                VALUES (?, ?, ?, ?)
            """, ("Cashier1", hashed_pw, "Nkiru", 1))
            app.logger.info("Billing user 'Nkiru' created successfully")
        else:
            app.logger.info("Billing user 'Nkiru' already exists")
            
    except Exception as e:
        app.logger.error(f"Error creating Nkiru user: {e}")
    
    conn.commit()
    cursor.close()
    conn.close()
    
def create_nkiru_user():
    """Create Nkiru as a default cashier/billing user."""
    conn = get_db_connection()
    if not conn:
        return
    
    cursor = conn.cursor()
    
    try:
        # Check if user already exists in cashier_users
        cursor.execute("SELECT id FROM cashier_users WHERE username = ?", ("Cashier1",))
        if not cursor.fetchone():
            hashed_pw = generate_password_hash("Nkiru1@allsaints")
            cursor.execute("""
                INSERT INTO cashier_users (username, password, full_name, is_active)
                VALUES (?, ?, ?, ?)
            """, ("Cashier1", hashed_pw, "Nkiru", 1))
            app.logger.info("Cashier user 'Nkiru' created successfully")
        else:
            app.logger.info("Cashier user 'Nkiru' already exists")
        
        # Check if user already exists in billing_users
        cursor.execute("SELECT id FROM billing_users WHERE username = ?", ("Cashier1",))
        if not cursor.fetchone():
            hashed_pw = generate_password_hash("Nkiru1@allsaints")
            cursor.execute("""
                INSERT INTO billing_users (username, password, full_name, is_active)
                VALUES (?, ?, ?, ?)
            """, ("Cashier1", hashed_pw, "Nkiru", 1))
            app.logger.info("Billing user 'Nkiru' created successfully")
        else:
            app.logger.info("Billing user 'Nkiru' already exists")
            
    except Exception as e:
        app.logger.error(f"Error creating Nkiru user: {e}")
    
    conn.commit()
    cursor.close()
    conn.close()


def create_christy_user():
    """Create Christy as a default pharmacy user."""
    conn = get_db_connection()
    if not conn:
        return
    
    cursor = conn.cursor()
    
    try:
        # Check if user already exists in pharmacists
        cursor.execute("SELECT id FROM pharmacists WHERE username = ?", ("Pharmacy1",))
        if not cursor.fetchone():
            hashed_pw = generate_password_hash("Christy1@allsaints")
            cursor.execute("""
                INSERT INTO pharmacists (username, password, full_name, is_active)
                VALUES (?, ?, ?, ?)
            """, ("Pharmacy1", hashed_pw, "Christy", 1))
            app.logger.info("Pharmacist user 'Christy' created successfully")
        else:
            app.logger.info("Pharmacist user 'Christy' already exists")
            
    except Exception as e:
        app.logger.error(f"Error creating Christy user: {e}")
    
    conn.commit()
    cursor.close()
    conn.close()
def create_default_admin():
    """Create default admin user."""
    conn = get_db_connection()
    if not conn:
        return

    cursor = conn.cursor()
    try:
        cursor.execute("SELECT COUNT(*) FROM admin_users WHERE username = 'admin'")
        if cursor.fetchone()[0] == 0:
            hashed_pw = generate_password_hash('admin123')
            cursor.execute("""
                INSERT INTO admin_users (username, password, full_name, email, role, is_super_admin)
                VALUES (?, ?, ?, ?, ?, ?)
            """, ('admin', hashed_pw, 'System Administrator', 'admin@hospital.com', 'Super Admin', 1))
            conn.commit()
            app.logger.info("Default admin user created")
    except Exception as e:
        app.logger.error(f"Error creating default admin: {e}")
    finally:
        cursor.close()
        conn.close()

def create_hr_tables():
    """Create HR-related tables in SQLite."""
    conn = get_db_connection()
    if not conn:
        app.logger.error("Cannot connect to database for HR table creation")
        return

    cursor = conn.cursor()
    
    # HR Users Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS hr_users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username VARCHAR(50) UNIQUE NOT NULL,
            password VARCHAR(255) NOT NULL,
            full_name VARCHAR(100) NOT NULL,
            email VARCHAR(100),
            role VARCHAR(50) DEFAULT 'HR Staff',
            is_active BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Departments Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS departments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name VARCHAR(100) NOT NULL,
            code VARCHAR(20) UNIQUE NOT NULL,
            description TEXT,
            head_of_dept VARCHAR(100),
            status VARCHAR(20) DEFAULT 'Active',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Staff Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS staff (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            staff_id VARCHAR(50) UNIQUE NOT NULL,
            first_name VARCHAR(100) NOT NULL,
            last_name VARCHAR(100) NOT NULL,
            department_id INTEGER REFERENCES departments(id),
            position VARCHAR(100) NOT NULL,
            employment_type VARCHAR(50),
            email VARCHAR(100),
            phone VARCHAR(20),
            hire_date DATE NOT NULL,
            salary DECIMAL(12, 2),
            status VARCHAR(20) DEFAULT 'Active',
            emergency_contact VARCHAR(100),
            address TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Attendance Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            staff_id INTEGER REFERENCES staff(id),
            date DATE NOT NULL,
            check_in TIME,
            check_out TIME,
            status VARCHAR(20),
            remarks TEXT,
            recorded_by INTEGER REFERENCES hr_users(id),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Leaves Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS leaves (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            staff_id INTEGER REFERENCES staff(id),
            leave_type VARCHAR(50) NOT NULL,
            start_date DATE NOT NULL,
            end_date DATE NOT NULL,
            days_requested INTEGER NOT NULL,
            reason TEXT,
            status VARCHAR(20) DEFAULT 'Pending',
            approved_by INTEGER REFERENCES hr_users(id),
            approved_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Schedules Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS schedules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            staff_id INTEGER REFERENCES staff(id),
            schedule_date DATE NOT NULL,
            shift_type VARCHAR(50),
            start_time TIME NOT NULL,
            end_time TIME NOT NULL,
            location VARCHAR(100),
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Payroll Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS payroll (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            staff_id INTEGER REFERENCES staff(id),
            pay_period VARCHAR(50),
            basic_salary DECIMAL(12, 2),
            allowances DECIMAL(12, 2),
            deductions DECIMAL(12, 2),
            net_salary DECIMAL(12, 2),
            status VARCHAR(20) DEFAULT 'Pending',
            payment_date DATE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Documents Table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            staff_id INTEGER REFERENCES staff(id),
            document_type VARCHAR(50),
            document_name VARCHAR(255),
            file_path VARCHAR(500),
            uploaded_by INTEGER REFERENCES hr_users(id),
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    
    # Create indexes
    indexes = [
        "CREATE INDEX IF NOT EXISTS idx_staff_department ON staff(department_id);",
        "CREATE INDEX IF NOT EXISTS idx_attendance_staff_date ON attendance(staff_id, date);",
        "CREATE INDEX IF NOT EXISTS idx_leaves_staff_status ON leaves(staff_id, status);",
        "CREATE INDEX IF NOT EXISTS idx_schedules_staff_date ON schedules(staff_id, schedule_date);",
        "CREATE INDEX IF NOT EXISTS idx_payroll_staff_period ON payroll(staff_id, pay_period);",
        "CREATE INDEX IF NOT EXISTS idx_staff_status ON staff(status);",
        "CREATE INDEX IF NOT EXISTS idx_attendance_date ON attendance(date);",
        "CREATE INDEX IF NOT EXISTS idx_leaves_status ON leaves(status);"
    ]
    
    for index_query in indexes:
        try:
            cursor.execute(index_query)
        except Exception as e:
            app.logger.warning(f"Could not create index: {e}")
    
    conn.commit()
    cursor.close()
    conn.close()
    
    # Insert default data
    create_default_hr_data()

def create_default_hr_data():
    """Insert default HR data into SQLite tables."""
    conn = get_db_connection()
    if not conn:
        return
    
    cursor = conn.cursor()
    
    try:
        # Default hashed password for 'hr@admin123'
        hashed_password = generate_password_hash('hr@admin123')
        
        # Insert default HR users
        cursor.execute("""
            INSERT OR IGNORE INTO hr_users (username, password, full_name, email, role) 
            VALUES 
                (?, ?, ?, ?, ?),
                (?, ?, ?, ?, ?)
        """, (
            'hr_admin', hashed_password, 'HR Administrator', 'admin@hospital.com', 'HR Manager',
            'hr_staff', hashed_password, 'HR Staff', 'staff@hospital.com', 'HR Officer'
        ))
        
        # Insert sample departments
        departments = [
            ('Administration', 'ADMIN', 'Hospital Administration and Management', 'Dr. John Smith'),
            ('Medical', 'MED', 'Medical Services Department', 'Dr. Sarah Johnson'),
            ('Nursing', 'NURS', 'Nursing Services', 'Mrs. Grace Williams'),
            ('Pharmacy', 'PHARM', 'Pharmacy Department', 'Mr. Michael Brown'),
            ('Laboratory', 'LAB', 'Laboratory Services', 'Dr. David Miller'),
            ('Radiology', 'RAD', 'Radiology Department', 'Dr. Lisa Davis'),
            ('Finance', 'FIN', 'Finance and Billing Department', 'Mr. Robert Wilson'),
            ('Human Resources', 'HR', 'Human Resources Department', 'Ms. Patricia Taylor'),
            ('Maintenance', 'MAINT', 'Facility Maintenance', 'Mr. Thomas Anderson'),
            ('Security', 'SEC', 'Hospital Security', 'Mr. Richard Clark')
        ]
        
        for dept in departments:
            cursor.execute("""
                INSERT OR IGNORE INTO departments (name, code, description, head_of_dept) 
                VALUES (?, ?, ?, ?)
            """, dept)
        
        conn.commit()
        
    except Exception as e:
        app.logger.error(f"Error in create_default_hr_data: {e}")
        conn.rollback()
    
    finally:
        cursor.close()
        conn.close()

def add_missing_columns():
    """Add missing columns to existing tables."""
    conn = get_db_connection()
    if not conn:
        return
    
    cursor = conn.cursor()
    
    # SQLite doesn't support adding multiple columns in one ALTER TABLE
    # We'll try each column individually
    columns_to_add = [
        ("billing_users", "created_by", "INTEGER"),
        ("pharmacists", "created_by", "INTEGER"),
        ("pharmacists", "full_name", "VARCHAR(100)"),
        ("billing_users", "full_name", "VARCHAR(100)"),
        ("receipts", "pharmacist", "VARCHAR(50)")
    ]
    
    for table, column, col_type in columns_to_add:
        try:
            cursor.execute(f"ALTER TABLE {table} ADD COLUMN {column} {col_type}")
            app.logger.info(f"Added column {column} to {table}")
        except sqlite3.OperationalError as e:
            if "duplicate column name" in str(e).lower():
                app.logger.info(f"Column {column} already exists in {table}")
            else:
                app.logger.warning(f"Could not add column {column} to {table}: {e}")
        except Exception as e:
            app.logger.warning(f"Could not add column {column} to {table}: {e}")
    
    # Update existing records
    try:
        cursor.execute("UPDATE billing_users SET full_name = username WHERE full_name IS NULL;")
        cursor.execute("UPDATE pharmacists SET full_name = username WHERE full_name IS NULL;")
        conn.commit()
        app.logger.info("Updated missing column values")
    except Exception as e:
        app.logger.warning(f"Could not update column values: {e}")
    
    cursor.close()
    conn.close()

def sync_existing_users():
    """Sync existing users from billing_users and pharmacists to new tables."""
    conn = get_db_connection()
    if not conn:
        return
    
    cursor = conn.cursor()
    
    try:
        # Sync existing billing users to cashier_users
        cursor.execute("""
            INSERT OR IGNORE INTO cashier_users (username, password, full_name, created_at)
            SELECT bu.username, bu.password, 
                   COALESCE(bu.full_name, bu.username) as full_name, 
                   bu.created_at
            FROM billing_users bu
            WHERE NOT EXISTS (
                SELECT 1 FROM cashier_users cu WHERE cu.username = bu.username
            )
        """)
        
        # Update pharmacists full_name
        cursor.execute("""
            UPDATE pharmacists 
            SET full_name = username 
            WHERE full_name IS NULL
        """)
        
        conn.commit()
        app.logger.info("Successfully synced existing users to new tables")
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error syncing existing users: {e}")
    
    finally:
        cursor.close()
        conn.close()

def log_admin_action(admin_id, action, details=None):
    """Log admin actions for audit trail."""
    conn = get_db_connection()
    if not conn:
        return
    
    cursor = conn.cursor()
    try:
        # Get request context if available
        ip_address = request.remote_addr if hasattr(request, 'remote_addr') else None
        user_agent = request.user_agent.string if hasattr(request, 'user_agent') else None
        
        cursor.execute("""
            INSERT INTO admin_audit_logs (admin_id, action, details, ip_address, user_agent)
            VALUES (?, ?, ?, ?, ?)
        """, (admin_id, action, details, ip_address, user_agent))
        conn.commit()
    except Exception as e:
        app.logger.error(f"Error logging admin action: {e}")
    finally:
        cursor.close()
        conn.close()

# -------------------- HELPER FUNCTIONS --------------------
def format_currency(amount):
    """Format amount as Nigerian Naira currency."""
    return f"₦{amount:,.2f}"

def build_stock_snapshot(rows, today):
    stock = []
    for r in rows:
        expiry_date = r[5]
        quantity = r[3]
        threshold = r[6] or 20

        if expiry_date:
            expiry_date_obj = datetime.strptime(expiry_date, '%Y-%m-%d').date() if isinstance(expiry_date, str) else expiry_date
            days_left = (expiry_date_obj - today).days
            status = "EXPIRED" if days_left < 0 else "EXPIRING_SOON" if days_left <= 30 else "VALID"
        else:
            days_left = None
            status = "UNKNOWN"

        stock.append({
            "id": r[0], "name": r[1], "strength": r[2],
            "quantity": quantity, "unit_price": r[4],
            "expiry_date": expiry_date, "days_left": days_left,
            "status": status, "low_stock_threshold": threshold,
            "total_value": quantity * r[4]
        })
    return stock

def apply_stock_filter(stock, filter_type):
    if filter_type == "expired":
        return [d for d in stock if d["status"] == "EXPIRED"]
    elif filter_type in ("expiring", "expiring_soon"):
        return [d for d in stock if d["status"] == "EXPIRING_SOON"]
    elif filter_type in ("low", "low_stock"):
        return [d for d in stock if d["quantity"] <= d["low_stock_threshold"]]
    return stock

# -------------------- ROUTES: LANDING & MODULES --------------------
@app.route('/')
def landing_page():
    hospital_name = "All Saint Medical Center Nsukka, Enugu State"
    modules = [
        "System Admin", "Patient Services", "Clinical Services",
        "Pharmacy", "Laboratory", "Radiology", "Billing and Revenue",
        "Human Resources", "Management and Reports"
    ]
    return render_template("dashboard.html", hospital_name=hospital_name, modules=modules)

@app.route('/<module_name>')
def module_placeholder(module_name):
    display_name = module_name.replace('_', ' ').title()
    if module_name.lower() == "pharmacy":
        return redirect(url_for('pharmacy_login'))
    return render_template("module_placeholder.html", module_name=display_name)

# -------------------- ROUTES: PHARMACY MODULE --------------------
@app.route('/pharmacy/login', methods=['GET', 'POST'])
def pharmacy_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        if not conn:
            flash("Database connection error", "danger")
            return render_template("pharmacy_login.html")

        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, username, password FROM pharmacists WHERE username=? AND is_active=1",
            (username,)
        )
        pharmacist = cursor.fetchone()
        cursor.close()
        conn.close()

        if pharmacist and check_password_hash(pharmacist[2], password):
            session['pharmacist_id'] = pharmacist[0]
            session['pharmacist_username'] = pharmacist[1]
            return redirect(url_for('pharmacy_dashboard'))
        else:
            flash("Invalid username or password", "danger")

    return render_template("pharmacy_login.html")

@app.route('/pharmacy/dashboard')
def pharmacy_dashboard():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))
    return render_template(
        "pharmacy_dashboard.html",
        pharmacist_name=session.get('pharmacist_username')
    )

@app.route('/pharmacy/logout')
def pharmacy_logout():
    session.clear()
    return redirect(url_for('pharmacy_login'))

@app.route('/pharmacy/drug_sales')
def drug_sales():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))
    return render_template(
        "drug_sales_dashboard.html",
        pharmacist_name=session.get('pharmacist_username'),
        hospital_name="All Saint Medical Center Nsukka, Enugu State"
    )

@app.route('/pharmacy/add-stock', methods=['GET', 'POST'])
def add_stock():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    if request.method == 'POST':
        drug_name = request.form.get('drug_name', '').strip()
        strength = request.form.get('strength', '').strip()
        unit_price = request.form.get('unit_price', '').strip()
        quantity = request.form.get('quantity', '').strip()
        expiry_date = request.form.get('expiry_date', '').strip()

        if not all([drug_name, strength, unit_price, quantity, expiry_date]):
            flash("All fields including expiry date are required.", "danger")
            return redirect(url_for('add_stock'))

        try:
            unit_price = float(unit_price)
            quantity = int(quantity)
            expiry_date_obj = datetime.strptime(expiry_date, "%Y-%m-%d").date()
        except ValueError:
            flash("Invalid input data.", "danger")
            return redirect(url_for('add_stock'))

        conn = get_db_connection()
        if not conn:
            flash("Database connection error.", "danger")
            return redirect(url_for('add_stock'))

        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, stock_quantity FROM drugs
            WHERE name = ? AND strength = ? AND expiry_date = ?
        """, (drug_name, strength, expiry_date_obj.strftime('%Y-%m-%d')))

        existing = cursor.fetchone()

        if existing:
            cursor.execute("""
                UPDATE drugs
                SET stock_quantity = stock_quantity + ?,
                    unit_price = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (quantity, unit_price, existing[0]))
        else:
            cursor.execute("""
                INSERT INTO drugs (name, strength, unit_price, stock_quantity, expiry_date)
                VALUES (?, ?, ?, ?, ?)
            """, (drug_name, strength, unit_price, quantity, expiry_date_obj.strftime('%Y-%m-%d')))

        conn.commit()
        cursor.close()
        conn.close()

        flash("Stock added successfully.", "success")
        return redirect(url_for('add_stock'))

    return render_template("add_stock.html")


@app.template_filter('format_date')
def format_date_filter(date_value, format='%Y-%m-%d'):
    """Safely format a date, handling None values."""
    if date_value is None:
        return 'Never'
    if isinstance(date_value, str):
        if date_value == 'Never':
            return date_value
        try:
            date_value = datetime.strptime(date_value, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            try:
                date_value = datetime.strptime(date_value, '%Y-%m-%d')
            except ValueError:
                return 'Invalid date'
    if hasattr(date_value, 'strftime'):
        return date_value.strftime(format)
    return str(date_value)
@app.route('/api/drugs')
def api_drugs():
    if 'pharmacist_id' not in session:
        return jsonify([])

    search = request.args.get('q', '').strip()
    conn = get_db_connection()
    if not conn:
        return jsonify([])

    cur = conn.cursor()
    query = """
        SELECT id, name, strength, unit_price, stock_quantity
        FROM drugs
        WHERE stock_quantity > 0
    """
    params = ()
    
    if search:
        query += " AND LOWER(name) LIKE LOWER(?)"
        params = (f"{search}%",)
    
    query += " ORDER BY name ASC"
    cur.execute(query, params)
    
    rows = cur.fetchall()
    cur.close()
    conn.close()

    return jsonify([{
        "id": r[0], "name": r[1], "strength": r[2],
        "unit_price": float(r[3]), "stock_quantity": r[4]
    } for r in rows])

@app.route('/pharmacy/receipt', methods=['POST'])
def pharmacy_receipt():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    data = request.json
    receipt_no = f"RX-{uuid.uuid4().hex[:8].upper()}"

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO drug_sales
        (receipt_no, items, subtotal, discount, tax, grand_total, pharmacist)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        receipt_no,
        json.dumps(data["items"]),
        data["subtotal"],
        data["discount"],
        data["tax"],
        data["grand_total"],
        session.get('pharmacist_username')
    ))

    conn.commit()
    cur.close()
    conn.close()

    data["receipt_no"] = receipt_no
    return render_template(
        "receipt.html",
        receipt=data,
        hospital_name="All Saint Medical Center Nsukka, Enugu State",
        pharmacist_name=session.get('pharmacist_username')
    )

@app.route('/pharmacy/save-patient', methods=['POST'])
def save_patient_info():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    receipt_no = request.form['receipt_no']
    patient_name = request.form['patient_name']
    patient_id = request.form.get('patient_id')

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE drug_sales
        SET patient_name=?, patient_id=?
        WHERE receipt_no=?
    """, (patient_name, patient_id, receipt_no))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for('reprint_receipt', receipt_no=receipt_no))

@app.route('/pharmacy/receipt/<receipt_no>')
def reprint_receipt(receipt_no):
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT receipt_no, patient_name, patient_id, items,
               subtotal, discount, tax, grand_total, pharmacist, created_at
        FROM drug_sales
        WHERE receipt_no = ?
    """, (receipt_no,))

    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        flash("Receipt not found", "danger")
        return redirect(url_for('pharmacy_dashboard'))

    receipt = {
        "receipt_no": row[0], "patient_name": row[1], "patient_id": row[2],
        "items": json.loads(row[3]) if row[3] else [],
        "subtotal": float(row[4]), "discount": float(row[5]),
        "tax": float(row[6]), "grand_total": float(row[7]),
        "pharmacist": row[8], "date": row[9]
    }

    return render_template("receipt.html", receipt=receipt, hospital_name="All Saint Medical Center Nsukka, Enugu State")

@app.route("/pharmacy/confirm-payment", methods=["POST"])
def confirm_payment():
    if "pharmacist_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    data = request.get_json()
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO receipts (
                patient_name, patient_id, subtotal, discount, tax,
                total_amount, grand_total, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            RETURNING id;
        """, (
            data.get("patient_name"), data.get("patient_id"),
            data["subtotal"], data["discount"], data["tax"],
            data["grand_total"], data["grand_total"], datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ))

        receipt_id = cur.lastrowid

        for item in data["items"]:
            cur.execute("""
                INSERT INTO receipt_items (
                    receipt_id, drug_name, strength, quantity, unit_price
                )
                VALUES (?, ?, ?, ?, ?);
            """, (
                receipt_id, item["drug_name"], item["strength"],
                item["quantity"], item["unit_price"]
            ))

            cur.execute("""
                UPDATE drugs
                SET stock_quantity = stock_quantity - ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE name = ? AND strength = ?;
            """, (
                item["quantity"], item["drug_name"], item["strength"]
            ))

        conn.commit()
        return jsonify({"success": True, "receipt_id": receipt_id})

    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

    finally:
        cur.close()
        conn.close()

@app.route("/pharmacy/stock-report")
def stock_report():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    filter_type = request.args.get("filter", "all")
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, strength, stock_quantity, unit_price, 
               expiry_date, low_stock_threshold
        FROM drugs
        ORDER BY expiry_date ASC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    stock = build_stock_snapshot(rows, date.today())
    stock = apply_stock_filter(stock, filter_type)

    return render_template(
        "stock_report.html",
        stock=stock,
        current_filter=filter_type,
        expired_count=sum(1 for d in stock if d["status"] == "EXPIRED"),
        expiring_soon_count=sum(1 for d in stock if d["status"] == "EXPIRING_SOON"),
        low_stock_count=sum(1 for d in stock if d["quantity"] <= d["low_stock_threshold"]),
        total_stock_value=sum(d["total_value"] for d in stock)
    )

@app.route("/pharmacy/stock-report/export")
def export_stock_report():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    filter_type = request.args.get("filter", "all")
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, strength, stock_quantity, unit_price, 
               expiry_date, low_stock_threshold
        FROM drugs
        ORDER BY expiry_date ASC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    stock = build_stock_snapshot(rows, date.today())
    stock = apply_stock_filter(stock, filter_type)

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Report"

    headers = [
        "Drug Name", "Strength", "Quantity", "Unit Price (₦)",
        "Expiry Date", "Days Left", "Status", "Total Value (₦)", "Low Stock Threshold"
    ]
    ws.append(headers)

    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    fills = {
        "EXPIRED": PatternFill("solid", fgColor="FF9999"),
        "EXPIRING_SOON": PatternFill("solid", fgColor="FFFF99"),
        "LOW": PatternFill("solid", fgColor="ADD8E6")
    }

    for item in stock:
        ws.append([
            item["name"], item["strength"], item["quantity"],
            float(item["unit_price"]), item["expiry_date"],
            item["days_left"], item["status"],
            float(item["total_value"]), item["low_stock_threshold"]
        ])

        row_idx = ws.max_row
        if item["status"] in fills:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fills[item["status"]]
        elif item["quantity"] <= item["low_stock_threshold"]:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fills["LOW"]

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name=f"pharmacy_stock_report_{date.today()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/pharmacy/stock-movements')
def stock_movements():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT sm.id, d.name, d.strength, sm.movement_type, 
               sm.quantity, u.username, sm.created_at, sm.note
        FROM stock_movements sm
        JOIN drugs d ON sm.drug_id = d.id
        JOIN users u ON sm.user_id = u.id
        ORDER BY sm.created_at DESC
    """)
    movements = cur.fetchall()
    cur.close()
    conn.close()

    return render_template('stock_movements.html', movements=movements)

@app.route("/pharmacy/revenue-report", methods=["GET", "POST"])
def revenue_report():
    if 'pharmacist_id' not in session:
        return redirect(url_for('pharmacy_login'))

    report_type = request.form.get("period", "daily")
    selected_day = request.form.get("day")
    selected_month = request.form.get("month")
    selected_year = request.form.get("year")
    today = date.today()

    if report_type == "daily":
        start_date = end_date = datetime.strptime(selected_day, "%Y-%m-%d").date() if selected_day else today
    elif report_type == "weekly":
        d = datetime.strptime(selected_day, "%Y-%m-%d").date() if selected_day else today
        start_date = d - timedelta(days=d.weekday())
        end_date = start_date + timedelta(days=6)
    elif report_type == "monthly":
        month = int(selected_month) if selected_month else today.month
        year = int(selected_year) if selected_year else today.year
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
    else:
        flash("Invalid report period", "danger")
        return redirect(url_for("pharmacy_dashboard"))

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, patient_name, patient_id, grand_total, created_at
        FROM receipts
        WHERE DATE(created_at) BETWEEN ? AND ?
        ORDER BY created_at ASC;
    """, (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
    sales = cur.fetchall()
    cur.close()
    conn.close()

    total_revenue = sum(float(s[3]) if s[3] is not None else 0.0 for s in sales)
    months = [(i, month_name[i]) for i in range(1, 13)]
    years = range(2024, today.year + 1)

    return render_template(
        "revenue_report.html",
        sales=sales,
        total_revenue=total_revenue,
        period=report_type,
        start_date=start_date,
        end_date=end_date,
        selected_day=selected_day,
        selected_month=int(selected_month) if selected_month else today.month,
        selected_year=int(selected_year) if selected_year else today.year,
        months=months,
        years=years
    )

@app.route("/receipt/<int:receipt_id>")
def receipt(receipt_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM receipts WHERE id = ?;", (receipt_id,))
    receipt = cur.fetchone()

    cur.execute("""
        SELECT drug_name, strength, quantity, unit_price
        FROM receipt_items
        WHERE receipt_id = ?;
    """, (receipt_id,))
    items = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "receipt.html",
        receipt=receipt,
        items=items
    )

@app.route("/pharmacy/receipt/<int:receipt_id>")
def view_receipt(receipt_id):
    if "pharmacist_id" not in session:
        return redirect(url_for("pharmacy_login"))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, patient_name, patient_id, subtotal, discount, tax, grand_total, created_at
        FROM receipts
        WHERE id = ?
    """, (receipt_id,))

    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        flash("Receipt not found", "danger")
        return redirect(url_for("pharmacy_dashboard"))

    receipt = {
        "id": row[0],
        "patient_name": row[1],
        "patient_id": row[2],
        "subtotal": float(row[3]),
        "discount": float(row[4]),
        "tax": float(row[5]),
        "grand_total": float(row[6]),
        "date": row[7]
    }

    cur.execute("""
        SELECT drug_name, strength, quantity, unit_price
        FROM receipt_items
        WHERE receipt_id = ?
    """, (receipt_id,))

    items_rows = cur.fetchall()
    items = []
    for i in items_rows:
        items.append({
            "drug_name": i[0],
            "strength": i[1],
            "quantity": i[2],
            "unit_price": float(i[3])
        })

    cur.close()
    conn.close()

    return render_template(
        "receipt.html",
        receipt=receipt,
        items=items,
        hospital_name="All Saint Medical Center, Nsukka, Enugu State"
    )

# -------------------- ROUTES: BILLING MODULE --------------------
@app.route("/billing/login", methods=["GET", "POST"])
def billing_login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT id, password
            FROM billing_users
            WHERE username = ?
        """, (username,))

        user = cur.fetchone()
        cur.close()
        conn.close()

        if user and check_password_hash(user[1], password):
            session["billing_user_id"] = user[0]
            session["billing_username"] = username
            return redirect(url_for("billing_dashboard"))
        else:
            flash("Invalid login credentials", "danger")

    return render_template("billing_login.html")

@app.route("/billing/dashboard")
def billing_dashboard():
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))
    return render_template("billing_dashboard.html")

@app.route("/billing/logout")
def billing_logout():
    session.pop("billing_user_id", None)
    session.pop("billing_username", None)
    flash("Logged out successfully", "success")
    return redirect(url_for("billing_login"))

@app.route("/billing/confirm-payment", methods=["POST"])
def billing_confirm_payment():
    if "billing_user_id" not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    patient_name = request.form.get("patient_name")
    service_type = request.form.get("service_type")
    payment_method = request.form.get("payment_method")
    amount_paid = float(request.form.get("amount_paid", 0))
    vat_percent = float(request.form.get("vat", 0))
    discount = float(request.form.get("discount", 0))

    subtotal = amount_paid
    vat_amount = (subtotal * vat_percent) / 100
    grand_total = subtotal + vat_amount - discount
    balance = 0 if amount_paid >= grand_total else grand_total - amount_paid
    status = "Paid" if balance <= 0 else "Partial"

    # Always use today's date — never trust the form's date field
    payment_date = date.today().strftime('%Y-%m-%d')

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO payments (
                patient_name, service_type, subtotal, discount, tax,
                grand_total, amount_paid, balance, payment_method,
                status, payment_date, recorded_by
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            patient_name, service_type, subtotal, discount, vat_amount,
            grand_total, amount_paid, balance, payment_method,
            status, payment_date, session["billing_user_id"]
        ))

        payment_id = cur.lastrowid
        conn.commit()
        flash(f"Payment recorded successfully. Receipt No: {payment_id}", "success")
        return redirect(url_for("view_payment_receipt", payment_id=payment_id))

    except Exception as e:
        conn.rollback()
        app.logger.error(f"Payment error: {e}")
        flash(f"Payment error: {str(e)}", "danger")
        return redirect(url_for("accept_payment_page"))

    finally:
        cur.close()
        conn.close()        
        
        
        
@app.route("/billing/accept-payment", methods=["GET"])
def accept_payment_page():
    return render_template("accept_payment.html")

@app.route("/billing/payment-history")
def payment_history():
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))

    patient_name = request.args.get("patient_name", "").strip()
    service_type = request.args.get("service_type", "")
    payment_method = request.args.get("payment_method", "")
    status = request.args.get("status", "")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    page = request.args.get("page", 1, type=int)
    per_page = 20

    conn = get_db_connection()
    cur = conn.cursor()

    query = "SELECT * FROM payments WHERE 1=1"
    count_query = "SELECT COUNT(*) FROM payments WHERE 1=1"
    params = []

    if patient_name:
        query += " AND LOWER(patient_name) LIKE LOWER(?)"
        count_query += " AND LOWER(patient_name) LIKE LOWER(?)"
        params.append(f"%{patient_name}%")

    if service_type:
        query += " AND service_type = ?"
        count_query += " AND service_type = ?"
        params.append(service_type)

    if payment_method:
        query += " AND payment_method = ?"
        count_query += " AND payment_method = ?"
        params.append(payment_method)

    if status:
        query += " AND status = ?"
        count_query += " AND status = ?"
        params.append(status)

    if start_date:
        query += " AND DATE(payment_date) >= ?"
        count_query += " AND DATE(payment_date) >= ?"
        params.append(start_date)

    if end_date:
        query += " AND DATE(payment_date) <= ?"
        count_query += " AND DATE(payment_date) <= ?"
        params.append(end_date)

    cur.execute(count_query, params)
    total_items = cur.fetchone()[0]

    query += " ORDER BY created_at DESC LIMIT ? OFFSET ?"
    offset = (page - 1) * per_page
    params.extend([per_page, offset])

    cur.execute(query, params)
    payments = cur.fetchall()

    cur.execute("SELECT DISTINCT service_type FROM payments WHERE service_type IS NOT NULL ORDER BY service_type")
    service_types = [row[0] for row in cur.fetchall()]

    def parse_dt(val):
        if not val or not isinstance(val, str):
            return val
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue
        return None

    total_amount = 0
    formatted_payments = []
    for payment in payments:
        payment_dict = dict(payment)
        payment_dict["amount_paid"] = float(payment_dict["amount_paid"])
        payment_dict["payment_date"] = parse_dt(payment_dict.get("payment_date"))
        payment_dict["created_at"] = parse_dt(payment_dict.get("created_at"))
        formatted_payments.append(payment_dict)
        total_amount += payment_dict["amount_paid"]

    cur.close()
    conn.close()

    total_pages = (total_items + per_page - 1) // per_page

    return render_template(
        "billing_payment_history.html",
        payments=formatted_payments,
        service_types=service_types,
        total_items=total_items,
        total_amount=total_amount,
        page=page,
        total_pages=total_pages,
        current_filters=request.args
    )
    
def parse_date(value):
    """Safely parse a date string or return None."""
    if not value or not isinstance(value, str):
        return value  # already a datetime, None, or numeric — caller handles it
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None

@app.route("/billing/payment-history/export")
def export_payment_history():
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))
    
    patient_name = request.args.get("patient_name", "").strip()
    service_type = request.args.get("service_type", "")
    payment_method = request.args.get("payment_method", "")
    status = request.args.get("status", "")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    query = "SELECT * FROM payments WHERE 1=1"
    params = []
    
    if patient_name:
        query += " AND LOWER(patient_name) LIKE LOWER(?)"
        params.append(f"%{patient_name}%")
    
    if service_type:
        query += " AND service_type = ?"
        params.append(service_type)
    
    if payment_method:
        query += " AND payment_method = ?"
        params.append(payment_method)
    
    if status:
        query += " AND status = ?"
        params.append(status)
    
    if start_date:
        query += " AND payment_date >= ?"
        params.append(start_date)
    
    if end_date:
        query += " AND payment_date <= ?"
        params.append(end_date)
    
    query += " ORDER BY created_at DESC"
    cur.execute(query, params)
    payments = cur.fetchall()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment History"
    
    headers = [
        "Receipt No", "Patient Name", "Service Type", 
        "Subtotal (₦)", "Discount (₦)", "Tax (₦)", "Grand Total (₦)",
        "Amount Paid (₦)", "Balance (₦)", "Payment Method",
        "Status", "Payment Date", "Created At", "Recorded By"
    ]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for payment in payments:
        ws.append([
            payment[0], payment[1], payment[2],
            float(payment[3]), float(payment[4]), float(payment[5]),
            float(payment[6]), float(payment[7]), float(payment[8]),
            payment[9], payment[10], payment[11], payment[13], payment[12]
        ])
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    ws.append([])
    ws.append(["SUMMARY", "", "", "", "", "", "", "", "", "", "", "", "", ""])
    
    if payments:
        total_amount = sum(float(p[7]) for p in payments)
        total_balance = sum(float(p[8]) for p in payments)
        
        ws.append(["Total Transactions", len(payments)])
        ws.append(["Total Amount", total_amount])
        ws.append(["Total Balance", total_balance])
    
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    cur.close()
    conn.close()
    
    filename = f"payment_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/billing/receipt/<int:payment_id>")
def view_payment_receipt(payment_id):
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM payments WHERE id = ?", (payment_id,))
    payment = cur.fetchone()
    cur.close()
    conn.close()

    if not payment:
        flash("Receipt not found", "danger")
        return redirect(url_for("billing_dashboard"))

    receipt = {
        "id": payment[0],
        "patient_name": payment[1],
        "service_type": payment[2],
        "subtotal": float(payment[3]),
        "discount": float(payment[4]),
        "tax": float(payment[5]),
        "grand_total": float(payment[6]),
        "amount_paid": float(payment[7]),
        "balance": float(payment[8]),
        "payment_method": payment[9],
        "status": payment[10],
        "payment_date": payment[11],
        "created_at": payment[13]
    }

    return render_template("billing_receipt.html", receipt=receipt, hospital_name="All Saint Medical Center Nsukka, Enugu State")

@app.route("/billing/todays-collection")
def todays_collection():
    if "billing_user_id" not in session:
        return redirect(url_for("billing_login"))
    
    today = date.today()
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT id, patient_name, service_type, amount_paid, 
               payment_method, status, created_at
        FROM payments 
        WHERE DATE(payment_date) = DATE('now', 'localtime')
        ORDER BY created_at DESC
    """)
    
    today_payments = cur.fetchall()
    
    payment_methods_data = {
        'Cash': {'amount': 0, 'count': 0},
        'Card': {'amount': 0, 'count': 0},
        'Transfer': {'amount': 0, 'count': 0},
        'POS': {'amount': 0, 'count': 0},
        'Insurance': {'amount': 0, 'count': 0},
        'Other': {'amount': 0, 'count': 0}
    }
    
    total_transactions = len(today_payments)
    grand_total = 0
    amounts = []
    
    def parse_dt(val):
        """Parse a date/datetime string into a datetime object."""
        if not val or not isinstance(val, str):
            return val
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue
        return None

    recent_transactions = []
    for payment in today_payments:
        amount_paid = float(payment[3])
        payment_method = payment[4]
        
        grand_total += amount_paid
        amounts.append(amount_paid)
        
        if payment_method in payment_methods_data:
            payment_methods_data[payment_method]['amount'] += amount_paid
            payment_methods_data[payment_method]['count'] += 1
        else:
            payment_methods_data['Other']['amount'] += amount_paid
            payment_methods_data['Other']['count'] += 1
        
        recent_transactions.append({
            'id': payment[0],
            'patient_name': payment[1],
            'service_type': payment[2],
            'amount_paid': amount_paid,
            'payment_method': payment_method,
            'status': payment[5],
            'created_at': parse_dt(payment[6])
        })
    
    average_transaction = grand_total / total_transactions if total_transactions > 0 else 0
    highest_transaction = max(amounts) if amounts else 0
    lowest_transaction = min(amounts) if amounts else 0
    
    morning_total = 0
    afternoon_total = 0
    evening_total = 0
    
    for payment in today_payments:
        created_at = parse_dt(payment[6])
        if created_at and hasattr(created_at, 'hour'):
            hour = created_at.hour
            amount = float(payment[3])
            
            if 6 <= hour < 12:
                morning_total += amount
            elif 12 <= hour < 16:
                afternoon_total += amount
            elif 16 <= hour < 22:
                evening_total += amount
    
    payment_methods = []
    for method_name, data in payment_methods_data.items():
        if data['count'] > 0:
            percentage = (data['amount'] / grand_total * 100) if grand_total > 0 else 0
            payment_methods.append({
                'name': method_name,
                'amount': data['amount'],
                'count': data['count'],
                'percentage': round(percentage, 1)
            })
    
    daily_target = 500000.00
    
    cur.close()
    conn.close()
    
    today_date = today.strftime("%A, %B %d, %Y")
    
    return render_template(
        "todays_collection.html",
        today_date=today_date,
        grand_total=grand_total,
        cash_total=payment_methods_data['Cash']['amount'],
        card_total=payment_methods_data['Card']['amount'],
        transfer_total=payment_methods_data['Transfer']['amount'],
        pos_total=payment_methods_data['POS']['amount'],
        insurance_total=payment_methods_data['Insurance']['amount'],
        other_total=payment_methods_data['Other']['amount'],
        payment_methods=payment_methods,
        recent_transactions=recent_transactions,
        total_transactions=total_transactions,
        average_transaction=average_transaction,
        highest_transaction=highest_transaction,
        lowest_transaction=lowest_transaction,
        morning_total=morning_total,
        afternoon_total=afternoon_total,
        evening_total=evening_total,
        daily_target=daily_target
    )
    
    
# Custom filter for currency formatting
@app.template_filter('currency')
def currency_filter(amount):
    if amount is None:
        return "₦0.00"
    return f"₦{float(amount):,.2f}"

# -------------------- ROUTES: HR MODULE --------------------
@app.route("/hr/login", methods=["GET", "POST"])
def hr_login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        conn = get_db_connection()
        if not conn:
            flash("Database connection error", "danger")
            return render_template("hr_login.html")

        cur = conn.cursor()

        try:
            cur.execute("""
                SELECT id, username, password, full_name, role
                FROM hr_users
                WHERE username = ? AND is_active = 1
            """, (username,))

            user = cur.fetchone()
            cur.close()
            conn.close()

            if user:
                stored_hash = user[2]
                known_hash = "$2b$12$LQv3c1yqBWVHxkd0LsZcdeJN8L7Fmm8Zz3qG9XwFk8kC1YdV6n4Oq"
                
                if stored_hash == known_hash and password == "hr@admin123":
                    session["hr_user_id"] = user[0]
                    session["hr_username"] = user[1]
                    session["hr_full_name"] = user[3]
                    session["hr_role"] = user[4]
                    flash(f"Welcome, {user[3]}!", "success")
                    return redirect(url_for("hr_dashboard"))
                else:
                    flash("Invalid password. Use: hr@admin123", "danger")
            else:
                flash("Invalid username. Use: hr_admin or hr_staff", "danger")

        except Exception as e:
            app.logger.error(f"HR login error: {e}")
            flash("Login error. Please try again.", "danger")

    return render_template("hr_login.html")

@app.route("/hr/dashboard")
def hr_dashboard():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    if not conn:
        flash("Database connection error", "danger")
        return redirect(url_for("hr_login"))
    
    cur = conn.cursor()
    
    try:
        cur.execute("SELECT COUNT(*) FROM staff WHERE status = 'Active'")
        total_staff = cur.fetchone()[0] or 0
        
        today = date.today().strftime('%Y-%m-%d')
        cur.execute("""
            SELECT COUNT(DISTINCT staff_id) 
            FROM attendance 
            WHERE date = ? AND status IN ('Present', 'Late')
        """, (today,))
        active_staff = cur.fetchone()[0] or 0
        
        cur.execute("""
            SELECT COUNT(*) 
            FROM leaves 
            WHERE ? BETWEEN start_date AND end_date 
            AND status = 'Approved'
        """, (today,))
        on_leave = cur.fetchone()[0] or 0
        
        cur.execute("SELECT COUNT(*) FROM departments WHERE status = 'Active'")
        departments_count = cur.fetchone()[0] or 0
        
        cur.execute("SELECT COUNT(*) FROM leaves WHERE status = 'Pending'")
        pending_leave = cur.fetchone()[0] or 0
        
        cur.execute("""
            SELECT COUNT(*) 
            FROM schedules 
            WHERE schedule_date >= date('now')
        """)
        upcoming_shifts = cur.fetchone()[0] or 0
        
        cur.execute("""
            SELECT COUNT(*) 
            FROM staff 
            WHERE emergency_contact IS NULL OR address IS NULL
        """)
        pending_updates = cur.fetchone()[0] or 0
        
        cur.execute("""
            SELECT COUNT(*) 
            FROM attendance 
            WHERE date = ? AND status = 'Late'
        """, (today,))
        late_arrivals = cur.fetchone()[0] or 0
        
    except Exception as e:
        app.logger.error(f"Error fetching HR stats: {e}")
        total_staff = active_staff = on_leave = departments_count = 0
        pending_leave = upcoming_shifts = pending_updates = late_arrivals = 0
    
    finally:
        cur.close()
        conn.close()
    
    return render_template(
        "hr_dashboard.html",
        hospital_name="All Saint Medical Center Nsukka, Enugu State",
        total_staff=total_staff,
        active_staff=active_staff,
        on_leave=on_leave,
        departments_count=departments_count,
        pending_leave=pending_leave,
        upcoming_shifts=upcoming_shifts,
        pending_updates=pending_updates,
        late_arrivals=late_arrivals,
        current_year=date.today().year
    )

@app.route("/hr/logout")
def hr_logout():
    session.pop("hr_user_id", None)
    session.pop("hr_username", None)
    session.pop("hr_full_name", None)
    session.pop("hr_role", None)
    flash("Logged out successfully", "success")
    return redirect(url_for("hr_login"))

@app.route("/hr/staff-management")
def staff_management():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    if not conn:
        flash("Database connection error", "danger")
        return redirect(url_for("hr_login"))
    
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT s.id, s.staff_id, s.first_name, s.last_name, 
                   s.position, s.employment_type, s.email, s.phone,
                   s.hire_date, s.salary, s.status, s.emergency_contact,
                   s.address, d.name as department_name
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            ORDER BY s.id DESC
            LIMIT 100
        """)
        staff_list = cur.fetchall()
        
        cur.execute("SELECT COUNT(*) FROM staff WHERE status = 'Active'")
        total_staff = cur.fetchone()[0] or 0
        
        cur.execute("SELECT COUNT(*) FROM staff WHERE status = 'Active' AND employment_type = 'Full-Time'")
        active_staff = cur.fetchone()[0] or 0
        
        cur.execute("SELECT COUNT(*) FROM staff WHERE employment_type = 'Contract'")
        on_contract = cur.fetchone()[0] or 0
        
        cur.execute("SELECT COUNT(DISTINCT department_id) FROM staff")
        departments_count = cur.fetchone()[0] or 0
        
    except Exception as e:
        app.logger.error(f"Error fetching staff data: {e}")
        staff_list = []
        total_staff = active_staff = on_contract = departments_count = 0
    
    finally:
        cur.close()
        conn.close()
    
    return render_template("staff_management.html", 
                         staff_list=staff_list,
                         total_staff=total_staff,
                         active_staff=active_staff,
                         on_contract=on_contract,
                         departments_count=departments_count,
                         hospital_name="All Saint Medical Center Nsukka, Enugu State",
                         current_year=date.today().year)

@app.route("/hr/staff/<int:staff_id>")
def view_staff(staff_id):
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT s.*, d.name as department_name, d.code as department_code
            FROM staff s
            LEFT JOIN departments d ON s.department_id = d.id
            WHERE s.id = ?
        """, (staff_id,))
        
        staff = cur.fetchone()
        
        if not staff:
            flash("Staff member not found", "danger")
            return redirect(url_for("staff_management"))
        
        staff_dict = dict(staff)
        staff_dict["salary"] = float(staff_dict["salary"]) if staff_dict["salary"] else 0
        
        hire_date = staff_dict["hire_date"]
        if isinstance(hire_date, str):
            hire_date = datetime.strptime(hire_date, '%Y-%m-%d').date()
        
        today = date.today()
        years = today.year - hire_date.year
        months = today.month - hire_date.month
        
        if months < 0:
            years -= 1
            months += 12
        
        employment_duration = f"{years} year(s), {months} month(s)"
        
    except Exception as e:
        app.logger.error(f"Error fetching staff details: {e}")
        flash("Error loading staff details", "danger")
        return redirect(url_for("staff_management"))
    
    finally:
        cur.close()
        conn.close()
    
    return render_template("view_staff.html", 
                         staff=staff_dict,
                         today=today,
                         employment_duration=employment_duration,
                         hospital_name="All Saint Medical Center Nsukka, Enugu State")

@app.route("/hr/staff/add", methods=["GET", "POST"])
def add_staff():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    if request.method == "POST":
        staff_id = request.form.get('staff_id')
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        department_id = request.form.get('department_id')
        position = request.form.get('position')
        employment_type = request.form.get('employment_type')
        email = request.form.get('email')
        phone = request.form.get('phone')
        hire_date = request.form.get('hire_date')
        salary = request.form.get('salary')
        emergency_contact = request.form.get('emergency_contact')
        address = request.form.get('address')
        
        if not all([staff_id, first_name, last_name, department_id, position, hire_date]):
            flash("Please fill in all required fields", "danger")
            return redirect(url_for("add_staff"))
        
        conn = get_db_connection()
        if not conn:
            flash("Database connection error", "danger")
            return redirect(url_for("add_staff"))
        
        cur = conn.cursor()
        
        try:
            cur.execute("SELECT id FROM staff WHERE staff_id = ?", (staff_id,))
            if cur.fetchone():
                flash(f"Staff ID '{staff_id}' already exists. Please use a different ID.", "danger")
                return redirect(url_for("add_staff"))
            
            salary_decimal = float(salary) if salary else 0.00
            
            cur.execute("""
                INSERT INTO staff (
                    staff_id, first_name, last_name, department_id, 
                    position, employment_type, email, phone, 
                    hire_date, salary, emergency_contact, address, status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Active')
            """, (
                staff_id, first_name, last_name, department_id,
                position, employment_type, email, phone,
                hire_date, salary_decimal, emergency_contact, address
            ))
            
            conn.commit()
            flash(f"Staff member {first_name} {last_name} (ID: {staff_id}) added successfully!", "success")
            
            cur.execute("SELECT id FROM staff WHERE staff_id = ?", (staff_id,))
            new_staff_id = cur.fetchone()[0]
            return redirect(url_for("view_staff", staff_id=new_staff_id))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error adding staff: {e}")
            flash(f"Error adding staff: {str(e)}", "danger")
            return redirect(url_for("add_staff"))
            
        finally:
            cur.close()
            conn.close()
    
    conn = get_db_connection()
    if not conn:
        flash("Database connection error", "danger")
        return redirect(url_for("staff_management"))
    
    cur = conn.cursor()
    
    try:
        cur.execute("SELECT id, name, code FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
        cur.execute("""
            SELECT MAX(staff_id) FROM staff 
            WHERE staff_id GLOB 'EMP[0-9]*'
        """)
        last_staff_id = cur.fetchone()[0]
        
        if last_staff_id:
            import re
            match = re.search(r'EMP(\d+)', last_staff_id)
            if match:
                next_num = int(match.group(1)) + 1
                suggested_id = f"EMP{next_num:03d}"
            else:
                suggested_id = "EMP001"
        else:
            suggested_id = "EMP001"
            
        today = date.today().strftime("%Y-%m-%d")
        
    except Exception as e:
        app.logger.error(f"Error loading form data: {e}")
        departments = []
        suggested_id = "EMP001"
        today = date.today().strftime("%Y-%m-%d")
        
    finally:
        cur.close()
        conn.close()
    
    return render_template("add_staff.html",
                         departments=departments,
                         suggested_id=suggested_id,
                         today=today,
                         hospital_name="All Saint Medical Center Nsukka, Enugu State")

@app.route("/hr/staff/edit/<int:staff_id>", methods=["GET", "POST"])
def edit_staff(staff_id):
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    if request.method == "POST":
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        department_id = request.form.get('department_id')
        position = request.form.get('position')
        employment_type = request.form.get('employment_type')
        email = request.form.get('email')
        phone = request.form.get('phone')
        salary = request.form.get('salary')
        status = request.form.get('status')
        emergency_contact = request.form.get('emergency_contact')
        address = request.form.get('address')
        
        try:
            cur.execute("""
                UPDATE staff SET
                    first_name = ?,
                    last_name = ?,
                    department_id = ?,
                    position = ?,
                    employment_type = ?,
                    email = ?,
                    phone = ?,
                    salary = ?,
                    status = ?,
                    emergency_contact = ?,
                    address = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                first_name, last_name, department_id,
                position, employment_type, email, phone,
                salary, status, emergency_contact, address,
                staff_id
            ))
            
            conn.commit()
            flash("Staff details updated successfully!", "success")
            return redirect(url_for("view_staff", staff_id=staff_id))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error updating staff: {e}")
            flash("Error updating staff details", "danger")
    
    try:
        cur.execute("SELECT * FROM staff WHERE id = ?", (staff_id,))
        staff = cur.fetchone()
        
        if not staff:
            flash("Staff member not found", "danger")
            return redirect(url_for("staff_management"))
        
        cur.execute("SELECT id, name FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
    except Exception as e:
        app.logger.error(f"Error loading staff for edit: {e}")
        flash("Error loading staff details", "danger")
        return redirect(url_for("staff_management"))
    
    finally:
        cur.close()
        conn.close()
    
    return render_template("edit_staff.html",
                         staff=staff,
                         departments=departments,
                         hospital_name="All Saint Medical Center Nsukka, Enugu State")

# -------------------- ROUTES: SCHEDULING MODULE --------------------
@app.route("/hr/scheduling")
def scheduling():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        current_month = date.today().replace(day=1)
        if current_month.month == 12:
            next_month = current_month.replace(year=current_month.year + 1, month=1)
        else:
            next_month = current_month.replace(month=current_month.month + 1)
        
        cur.execute("""
            SELECT s.*, st.first_name, st.last_name, st.position, d.name as department_name
            FROM schedules s
            JOIN staff st ON s.staff_id = st.id
            LEFT JOIN departments d ON st.department_id = d.id
            WHERE s.schedule_date >= ? AND s.schedule_date < ?
            ORDER BY s.schedule_date, s.start_time
        """, (current_month.strftime('%Y-%m-%d'), next_month.strftime('%Y-%m-%d')))
        
        schedules = cur.fetchall()
        
        cur.execute("SELECT COUNT(*) FROM schedules WHERE schedule_date >= date('now')")
        upcoming_shifts = cur.fetchone()[0] or 0
        
        cur.execute("""
            SELECT COUNT(DISTINCT staff_id) 
            FROM schedules 
            WHERE schedule_date >= date('now')
        """)
        staff_scheduled = cur.fetchone()[0] or 0
        
        cur.execute("SELECT id, name FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
        cur.execute("""
            SELECT id, first_name, last_name, position 
            FROM staff 
            WHERE status = 'Active' 
            ORDER BY first_name, last_name
        """)
        staff_list = cur.fetchall()
        
    except Exception as e:
        app.logger.error(f"Error fetching scheduling data: {e}")
        schedules = []
        upcoming_shifts = 0
        staff_scheduled = 0
        departments = []
        staff_list = []
    
    finally:
        cur.close()
        conn.close()
    
    return render_template("scheduling_dashboard.html",
                         schedules=schedules,
                         upcoming_shifts=upcoming_shifts,
                         staff_scheduled=staff_scheduled,
                         departments=departments,
                         staff_list=staff_list,
                         current_month=current_month.strftime("%B %Y"),
                         hospital_name="All Saint Medical Center Nsukka, Enugu State")

@app.route("/hr/scheduling/create", methods=["GET", "POST"])
def create_schedule():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    if request.method == "POST":
        staff_id = request.form.get("staff_id")
        schedule_date = request.form.get("schedule_date")
        shift_type = request.form.get("shift_type")
        start_time = request.form.get("start_time")
        end_time = request.form.get("end_time")
        location = request.form.get("location")
        notes = request.form.get("notes")
        
        if not all([staff_id, schedule_date, start_time, end_time]):
            flash("Please fill in all required fields", "danger")
            return redirect(url_for("create_schedule"))
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        try:
            cur.execute("""
                SELECT id FROM schedules 
                WHERE staff_id = ? AND schedule_date = ?
            """, (staff_id, schedule_date))
            
            if cur.fetchone():
                flash("This staff already has a schedule for this date", "warning")
                return redirect(url_for("create_schedule"))
            
            cur.execute("""
                INSERT INTO schedules (
                    staff_id, schedule_date, shift_type, 
                    start_time, end_time, location, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (staff_id, schedule_date, shift_type, start_time, end_time, location, notes))
            
            conn.commit()
            flash("Schedule created successfully!", "success")
            return redirect(url_for("scheduling"))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error creating schedule: {e}")
            flash(f"Error creating schedule: {str(e)}", "danger")
            return redirect(url_for("create_schedule"))
            
        finally:
            cur.close()
            conn.close()
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT id, first_name, last_name, position, 
                   (SELECT name FROM departments WHERE id = staff.department_id) as department
            FROM staff 
            WHERE status = 'Active' 
            ORDER BY first_name, last_name
        """)
        staff_list = cur.fetchall()
        
        tomorrow = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
        
    except Exception as e:
        app.logger.error(f"Error loading schedule form data: {e}")
        staff_list = []
        tomorrow = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
        
    finally:
        cur.close()
        conn.close()
    
    return render_template("create_schedule.html",
                         staff_list=staff_list,
                         tomorrow=tomorrow,
                         hospital_name="All Saint Medical Center Nsukka, Enugu State")

@app.route("/hr/scheduling/roster")
def view_roster():
    if "hr_user_id" not in session:
        return redirect(url_for("hr_login"))
    
    department_id = request.args.get("department_id", "")
    staff_id = request.args.get("staff_id", "")
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    
    if not start_date:
        today = date.today()
        start_date = (today - timedelta(days=today.weekday())).strftime("%Y-%m-%d")
    
    if not end_date:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = (start_date_obj + timedelta(days=6)).strftime("%Y-%m-%d")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        query = """
            SELECT 
                s.id as schedule_id,
                s.schedule_date,
                s.shift_type,
                s.start_time,
                s.end_time,
                s.location,
                s.notes,
                st.id as staff_id,
                st.first_name,
                st.last_name,
                st.position,
                d.name as department_name,
                d.id as department_id
            FROM schedules s
            JOIN staff st ON s.staff_id = st.id
            LEFT JOIN departments d ON st.department_id = d.id
            WHERE s.schedule_date BETWEEN ? AND ?
        """
        params = [start_date, end_date]
        
        if department_id:
            query += " AND st.department_id = ?"
            params.append(department_id)
        
        if staff_id:
            query += " AND s.staff_id = ?"
            params.append(staff_id)
        
        query += " ORDER BY s.schedule_date, d.name, st.first_name, s.start_time"
        
        cur.execute(query, params)
        schedules = cur.fetchall()
        
        cur.execute("SELECT id, name FROM departments WHERE status = 'Active' ORDER BY name")
        departments = cur.fetchall()
        
        cur.execute("""
            SELECT id, first_name, last_name 
            FROM staff 
            WHERE status = 'Active' 
            ORDER BY first_name, last_name
        """)
        staff_list = cur.fetchall()
        
        schedule_dict = {}
        for schedule in schedules:
            schedule_date = schedule[1]
            if schedule_date not in schedule_dict:
                schedule_dict[schedule_date] = []
            
            schedule_dict[schedule_date].append({
                'id': schedule[0],
                'date': schedule[1],
                'shift_type': schedule[2],
                'start_time': schedule[3],
                'end_time': schedule[4],
                'location': schedule[5],
                'notes': schedule[6],
                'staff_id': schedule[7],
                'first_name': schedule[8],
                'last_name': schedule[9],
                'position': schedule[10],
                'department': schedule[11]
            })
        
        total_shifts = len(schedules)
        unique_staff = len(set([s[7] for s in schedules]))
        unique_departments = len(set([s[11] for s in schedules if s[11]]))
        
    except Exception as e:
        app.logger.error(f"Error fetching roster: {e}")
        schedules = []
        departments = []
        staff_list = []
        schedule_dict = {}
        total_shifts = 0
        unique_staff = 0
        unique_departments = 0
    
    finally:
        cur.close()
        conn.close()
    
    return render_template("view_roster.html",
                         schedules=schedules,
                         schedule_dict=schedule_dict,
                         departments=departments,
                         staff_list=staff_list,
                         start_date=start_date,
                         end_date=end_date,
                         selected_department=department_id,
                         selected_staff=staff_id,
                         total_shifts=total_shifts,
                         unique_staff=unique_staff,
                         unique_departments=unique_departments,
                         hospital_name="All Saint Medical Center Nsukka, Enugu State")

# -------------------- ROUTES: ADMIN MODULE --------------------
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        if not conn:
            flash("Database connection error", "danger")
            return render_template("admin_login.html")

        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, username, password, full_name, role, is_active 
            FROM admin_users 
            WHERE username=?
        """, (username,))
        admin = cursor.fetchone()

        if admin and admin[5]:  # Check if is_active
            if check_password_hash(admin[2], password):
                session['admin_id'] = admin[0]
                session['admin_username'] = admin[1]
                session['admin_full_name'] = admin[3]
                session['admin_role'] = admin[4]
                
                cursor.execute("""
                    UPDATE admin_users 
                    SET last_login = CURRENT_TIMESTAMP 
                    WHERE id = ?
                """, (admin[0],))
                conn.commit()
                
                log_admin_action(admin[0], 'LOGIN', f'User {username} logged in')
                
                flash(f"Welcome, {admin[3]}!", "success")
                return redirect(url_for('admin_dashboard'))
            else:
                flash("Invalid password", "danger")
        else:
            flash("Invalid username or account inactive", "danger")

        cursor.close()
        conn.close()

    return render_template("admin_login.html")

@app.route('/admin/dashboard')
def admin_dashboard():
    if 'admin_id' not in session:
        return redirect(url_for('admin_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    stats = {}
    try:
        cursor.execute("SELECT COUNT(*) FROM admin_users WHERE is_active = 1")
        stats['admins_count'] = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM pharmacists WHERE is_active = 1")
        stats['pharmacists_count'] = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM billing_users WHERE is_active = 1")
        stats['cashiers_count'] = cursor.fetchone()[0]
        
        today = date.today().strftime('%Y-%m-%d')
        cursor.execute("SELECT COALESCE(SUM(amount_paid), 0) FROM payments WHERE DATE(payment_date) = ?", (today,))
        stats['todays_collection'] = float(cursor.fetchone()[0] or 0)
        
        cursor.execute("""
            SELECT COUNT(*) FROM drugs 
            WHERE stock_quantity <= low_stock_threshold 
            AND expiry_date > date('now')
        """)
        stats['low_stock_count'] = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT COUNT(*) FROM drugs 
            WHERE expiry_date < date('now') AND stock_quantity > 0
        """)
        stats['expired_stock_count'] = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT COUNT(*), COALESCE(SUM(grand_total), 0)
            FROM receipts 
            WHERE DATE(created_at) = ?
        """, (today,))
        result = cursor.fetchone()
        stats['pharmacy_sales_count'] = result[0]
        stats['pharmacy_revenue'] = float(result[1] or 0)
        
        cursor.execute("""
            SELECT a.username, l.action, l.created_at 
            FROM admin_audit_logs l
            JOIN admin_users a ON l.admin_id = a.id
            ORDER BY l.created_at DESC 
            LIMIT 10
        """)
        recent_actions = cursor.fetchall()
        
        # Convert string dates to datetime objects for the template
        stats['recent_actions'] = []
        for action in recent_actions:
            username = action[0]
            action_text = action[1]
            created_at = action[2]
            
            # Convert string to datetime if needed
            if isinstance(created_at, str):
                try:
                    created_at = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        created_at = datetime.strptime(created_at, '%Y-%m-%d')
                    except ValueError:
                        created_at = datetime.now()
            
            stats['recent_actions'].append((username, action_text, created_at))
        
    except Exception as e:
        app.logger.error(f"Error fetching admin stats: {e}")
        flash("Error loading dashboard statistics", "warning")
        stats = {}
    
    cursor.close()
    conn.close()
    
    return render_template(
        "admin_dashboard.html",
        stats=stats,
        hospital_name="All Saint Medical Center Nsukka, Enugu State",
        admin_name=session.get('admin_full_name', 'Admin')
    )
    
def convert_dates_in_row(row):
    """Convert date strings in a row to datetime objects."""
    converted = []
    for value in row:
        if isinstance(value, str):
            # Try to parse as datetime
            try:
                # Try full datetime format
                converted.append(datetime.strptime(value, '%Y-%m-%d %H:%M:%S'))
                continue
            except ValueError:
                try:
                    # Try date format
                    converted.append(datetime.strptime(value, '%Y-%m-%d').date())
                    continue
                except ValueError:
                    pass
        converted.append(value)
    return tuple(converted)
@app.route('/admin/logout')
def admin_logout():
    if 'admin_id' in session:
        log_admin_action(session['admin_id'], 'LOGOUT')
    session.clear()
    flash("Logged out successfully", "success")
    return redirect(url_for('admin_login'))

# -------------------- ADMIN USER MANAGEMENT --------------------
@app.route('/admin/users/admins')
def admin_manage_admins():
    if 'admin_id' not in session:
        return redirect(url_for('admin_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT id, username, full_name, email, role, 
                   is_super_admin, is_active, created_at, last_login
            FROM admin_users 
            ORDER BY created_at DESC
        """)
        admins_raw = cursor.fetchall()
        
        # Convert dates to datetime objects and handle None values
        admins = []
        for admin in admins_raw:
            admin_list = list(admin)
            
            # Convert created_at
            if admin_list[7] and isinstance(admin_list[7], str):
                try:
                    admin_list[7] = datetime.strptime(admin_list[7], '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        admin_list[7] = datetime.strptime(admin_list[7], '%Y-%m-%d')
                    except ValueError:
                        admin_list[7] = None
            elif admin_list[7] and isinstance(admin_list[7], (int, float)):
                admin_list[7] = datetime.fromtimestamp(admin_list[7])
            
            # Convert last_login (might be None)
            if admin_list[8] and isinstance(admin_list[8], str):
                try:
                    admin_list[8] = datetime.strptime(admin_list[8], '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        admin_list[8] = datetime.strptime(admin_list[8], '%Y-%m-%d')
                    except ValueError:
                        admin_list[8] = None
            elif admin_list[8] and isinstance(admin_list[8], (int, float)):
                admin_list[8] = datetime.fromtimestamp(admin_list[8])
            
            admins.append(tuple(admin_list))
            
    except Exception as e:
        app.logger.error(f"Error fetching admins: {e}")
        admins = []
        flash("Error loading admin users", "danger")
    
    cursor.close()
    conn.close()
    
    return render_template(
        "admin_manage_admins.html",
        admins=admins,
        admin_name=session.get('admin_full_name', 'Admin')
    )


def convert_date_fields(row, date_indices):
    """
    Convert date fields in a row from strings to datetime objects.
    
    Args:
        row: Tuple of values from database
        date_indices: List of indices that contain date values
    
    Returns:
        Tuple with converted date fields
    """
    row_list = list(row)
    for idx in date_indices:
        if idx < len(row_list) and row_list[idx] and isinstance(row_list[idx], str):
            try:
                row_list[idx] = datetime.strptime(row_list[idx], '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    row_list[idx] = datetime.strptime(row_list[idx], '%Y-%m-%d')
                except ValueError:
                    row_list[idx] = None
    return tuple(row_list)

@app.route('/admin/users/create-admin', methods=['GET', 'POST'])
def admin_create_admin():
    if 'admin_id' not in session:
        return redirect(url_for('admin_login'))
    
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        full_name = request.form['full_name']
        email = request.form.get('email', '')
        role = request.form.get('role', 'Admin')
        
        if not all([username, password, full_name]):
            flash("Username, password, and full name are required", "danger")
            return redirect(url_for('admin_create_admin'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute("SELECT id FROM admin_users WHERE username = ?", (username,))
            if cursor.fetchone():
                flash("Username already exists", "danger")
                return redirect(url_for('admin_create_admin'))
            
            hashed_pw = generate_password_hash(password)
            cursor.execute("""
                INSERT INTO admin_users (username, password, full_name, email, role, created_by)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (username, hashed_pw, full_name, email, role, session['admin_id']))
            
            conn.commit()
            
            log_admin_action(session['admin_id'], 'CREATE_ADMIN', 
                           f'Created admin account: {username} ({full_name})')
            
            flash(f"Admin account created for {full_name}", "success")
            return redirect(url_for('admin_manage_admins'))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error creating admin: {e}")
            flash(f"Error creating admin account: {str(e)}", "danger")
            
        finally:
            cursor.close()
            conn.close()
    
    return render_template(
        "admin_create_user.html",
        user_type="admin",
        admin_name=session.get('admin_full_name', 'Admin')
    )

@app.route('/admin/users/create-cashier', methods=['GET', 'POST'])
def admin_create_cashier():
    if 'admin_id' not in session:
        return redirect(url_for('admin_login'))
    
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        full_name = request.form['full_name']
        email = request.form.get('email', '')
        
        if not all([username, password, full_name]):
            flash("Username, password, and full name are required", "danger")
            return redirect(url_for('admin_create_cashier'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute("SELECT id FROM cashier_users WHERE username = ?", (username,))
            if cursor.fetchone():
                flash("Username already exists in cashier users", "danger")
                return redirect(url_for('admin_create_cashier'))
            
            cursor.execute("SELECT id FROM billing_users WHERE username = ?", (username,))
            if cursor.fetchone():
                flash("Username already exists in billing users", "danger")
                return redirect(url_for('admin_create_cashier'))
            
            hashed_pw = generate_password_hash(password)
            
            cursor.execute("""
                INSERT INTO cashier_users (username, password, full_name, email, created_by)
                VALUES (?, ?, ?, ?, ?)
            """, (username, hashed_pw, full_name, email, session['admin_id']))
            
            try:
                cursor.execute("""
                    INSERT INTO billing_users (username, password, full_name, created_by)
                    VALUES (?, ?, ?, ?)
                """, (username, hashed_pw, full_name, session['admin_id']))
            except Exception as e:
                app.logger.warning(f"Could not insert into billing_users: {e}")
                cursor.execute("""
                    INSERT INTO billing_users (username, password, full_name)
                    VALUES (?, ?, ?)
                """, (username, hashed_pw, full_name))
            
            conn.commit()
            
            log_admin_action(session['admin_id'], 'CREATE_CASHIER', 
                           f'Created cashier account: {username} ({full_name})')
            
            flash(f"Cashier account created for {full_name}", "success")
            return redirect(url_for('admin_manage_cashiers'))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error creating cashier: {e}")
            flash(f"Error creating cashier account: {str(e)}", "danger")
            
        finally:
            cursor.close()
            conn.close()
    
    return render_template(
        "admin_create_user.html",
        user_type="cashier",
        admin_name=session.get('admin_full_name', 'Admin')
    )

@app.route('/admin/users/create-pharmacist', methods=['GET', 'POST'])
def admin_create_pharmacist():
    if 'admin_id' not in session:
        return redirect(url_for('admin_login'))
    
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        full_name = request.form['full_name']
        
        if not all([username, password, full_name]):
            flash("Username, password, and full name are required", "danger")
            return redirect(url_for('admin_create_pharmacist'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute("SELECT id FROM pharmacists WHERE username = ?", (username,))
            if cursor.fetchone():
                flash("Username already exists", "danger")
                return redirect(url_for('admin_create_pharmacist'))
            
            hashed_pw = generate_password_hash(password)
            cursor.execute("""
                INSERT INTO pharmacists (username, password, full_name, created_by)
                VALUES (?, ?, ?, ?)
            """, (username, hashed_pw, full_name, session['admin_id']))
            
            conn.commit()
            
            log_admin_action(session['admin_id'], 'CREATE_PHARMACIST', 
                           f'Created pharmacist account: {username} ({full_name})')
            
            flash(f"Pharmacist account created for {full_name}", "success")
            return redirect(url_for('admin_manage_pharmacists'))
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error creating pharmacist: {e}")
            flash(f"Error creating pharmacist account: {str(e)}", "danger")
            
        finally:
            cursor.close()
            conn.close()
    
    return render_template(
        "admin_create_user.html",
        user_type="pharmacist",
        admin_name=session.get('admin_full_name', 'Admin')
    )

@app.route('/admin/users/cashiers')
def admin_manage_cashiers():
    if 'admin_id' not in session:
        return redirect(url_for('admin_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT id, username, full_name, is_active, created_at, last_login
            FROM cashier_users 
            ORDER BY created_at DESC
        """)
        cashiers_raw = cursor.fetchall()
        
        # Convert dates to datetime objects and handle None values
        cashiers = []
        for cashier in cashiers_raw:
            cashier_list = list(cashier)
            
            # Convert created_at
            if cashier_list[4] and isinstance(cashier_list[4], str):
                try:
                    cashier_list[4] = datetime.strptime(cashier_list[4], '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        cashier_list[4] = datetime.strptime(cashier_list[4], '%Y-%m-%d')
                    except ValueError:
                        cashier_list[4] = None
            elif cashier_list[4] and isinstance(cashier_list[4], (int, float)):
                # If it's a timestamp
                cashier_list[4] = datetime.fromtimestamp(cashier_list[4])
            
            # Convert last_login (might be None)
            if cashier_list[5] and isinstance(cashier_list[5], str):
                try:
                    cashier_list[5] = datetime.strptime(cashier_list[5], '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        cashier_list[5] = datetime.strptime(cashier_list[5], '%Y-%m-%d')
                    except ValueError:
                        cashier_list[5] = None
            elif cashier_list[5] and isinstance(cashier_list[5], (int, float)):
                cashier_list[5] = datetime.fromtimestamp(cashier_list[5])
            # If it's None, leave it as None
            
            cashiers.append(tuple(cashier_list))
            
    except Exception as e:
        app.logger.error(f"Error fetching cashiers: {e}")
        cashiers = []
        flash("Error loading cashier users", "danger")
    
    cursor.close()
    conn.close()
    
    return render_template(
        "admin_manage_cashiers.html",
        cashiers=cashiers,
        admin_name=session.get('admin_full_name', 'Admin')
    )
    
    
@app.route('/admin/users/pharmacists')
def admin_manage_pharmacists():
    if 'admin_id' not in session:
        return redirect(url_for('admin_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # AFTER
    try:
        cursor.execute("""
            SELECT id, username, full_name, is_active, created_at
            FROM pharmacists 
            ORDER BY created_at DESC
        """)
        pharmacists_raw = cursor.fetchall()

        pharmacists = []
        for pharmacist in pharmacists_raw:
            pharmacist_list = list(pharmacist)

            # Convert created_at (index 4) from string to datetime
            if pharmacist_list[4] and isinstance(pharmacist_list[4], str):
                try:
                    pharmacist_list[4] = datetime.strptime(pharmacist_list[4], '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        pharmacist_list[4] = datetime.strptime(pharmacist_list[4], '%Y-%m-%d')
                    except ValueError:
                        pharmacist_list[4] = None
            elif pharmacist_list[4] and isinstance(pharmacist_list[4], (int, float)):
                pharmacist_list[4] = datetime.fromtimestamp(pharmacist_list[4])

            pharmacists.append(tuple(pharmacist_list))

    except Exception as e:
        app.logger.error(f"Error fetching pharmacists: {e}")
        pharmacists = []
        flash("Error loading pharmacist users", "danger")
    
    cursor.close()
    conn.close()
    
    return render_template(
        "admin_manage_pharmacists.html",
        pharmacists=pharmacists,
        admin_name=session.get('admin_full_name', 'Admin')
    )

@app.route('/admin/users/toggle-status/<user_type>/<int:user_id>', methods=['POST'])
def admin_toggle_user_status(user_type, user_id):
    if 'admin_id' not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        if user_type == 'admin':
            table = 'admin_users'
        elif user_type == 'cashier':
            table = 'cashier_users'
        elif user_type == 'pharmacist':
            table = 'pharmacists'
        else:
            return jsonify({"success": False, "message": "Invalid user type"}), 400
        
        cursor.execute(f"SELECT is_active FROM {table} WHERE id = ?", (user_id,))
        result = cursor.fetchone()
        if not result:
            return jsonify({"success": False, "message": "User not found"}), 404
        
        new_status = 1 if not result[0] else 0
        
        cursor.execute(f"UPDATE {table} SET is_active = ? WHERE id = ?", (new_status, user_id))
        conn.commit()
        
        status_text = "activated" if new_status else "deactivated"
        log_admin_action(session['admin_id'], 'TOGGLE_USER_STATUS', 
                       f'{status_text} {user_type} user ID: {user_id}')
        
        return jsonify({
            "success": True, 
            "message": f"User {status_text} successfully",
            "new_status": bool(new_status)
        })
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error toggling user status: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
        
    finally:
        cursor.close()
        conn.close()

# -------------------- ADMIN REPORTS --------------------
@app.route('/admin/reports/pharmacy-stock')
def admin_pharmacy_stock():
    if 'admin_id' not in session:
        return redirect(url_for('admin_login'))
    
    filter_type = request.args.get("filter", "all")
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT id, name, strength, stock_quantity, unit_price, 
                   expiry_date, low_stock_threshold
            FROM drugs
            ORDER BY expiry_date ASC
        """)
        rows = cur.fetchall()
        
        stock = build_stock_snapshot(rows, date.today())
        stock = apply_stock_filter(stock, filter_type)
        
        total_stock_value = sum(d["total_value"] for d in stock)
        total_items = len(rows)
        filtered_items = len(stock)
        
    except Exception as e:
        app.logger.error(f"Error fetching pharmacy stock: {e}")
        stock = []
        total_stock_value = 0
        total_items = 0
        filtered_items = 0
        flash("Error loading pharmacy stock report", "danger")
    
    cur.close()
    conn.close()
    
    return render_template(
        "admin_pharmacy_stock.html",
        stock=stock,
        current_filter=filter_type,
        total_stock_value=total_stock_value,
        total_items=total_items,
        filtered_items=filtered_items,
        admin_name=session.get('admin_full_name', 'Admin')
    )

@app.route('/admin/reports/pharmacy-revenue', methods=['GET', 'POST'])
def admin_pharmacy_revenue():
    if 'admin_id' not in session:
        return redirect(url_for('admin_login'))
    
    report_type = request.form.get("period", "daily")
    selected_day = request.form.get("day")
    selected_month = request.form.get("month")
    selected_year = request.form.get("year")
    today = date.today()
    
    if report_type == "daily":
        start_date = end_date = datetime.strptime(selected_day, "%Y-%m-%d").date() if selected_day else today
    elif report_type == "weekly":
        d = datetime.strptime(selected_day, "%Y-%m-%d").date() if selected_day else today
        start_date = d - timedelta(days=d.weekday())
        end_date = start_date + timedelta(days=6)
    elif report_type == "monthly":
        month = int(selected_month) if selected_month else today.month
        year = int(selected_year) if selected_year else today.year
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
    else:
        flash("Invalid report period", "danger")
        return redirect(url_for("admin_dashboard"))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT r.id, r.patient_name, r.patient_id, r.grand_total, 
                   r.created_at, r.pharmacist
            FROM receipts r
            WHERE DATE(r.created_at) BETWEEN ? AND ?
            ORDER BY r.created_at DESC
        """, (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        
        sales = cur.fetchall()
        
        cur.execute("""
            SELECT 
                COUNT(*) as transactions,
                COALESCE(SUM(grand_total), 0) as total_revenue,
                AVG(grand_total) as avg_transaction,
                MIN(grand_total) as min_transaction,
                MAX(grand_total) as max_transaction
            FROM receipts
            WHERE DATE(created_at) BETWEEN ? AND ?
        """, (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        
        stats = cur.fetchone()
        total_revenue = float(stats[1]) if stats[1] else 0.0
        avg_transaction = float(stats[2]) if stats[2] else 0.0
        
        cur.execute("""
            SELECT 
                ri.drug_name,
                ri.strength,
                SUM(ri.quantity) as total_quantity,
                SUM(ri.quantity * ri.unit_price) as total_value
            FROM receipt_items ri
            JOIN receipts r ON ri.receipt_id = r.id
            WHERE DATE(r.created_at) BETWEEN ? AND ?
            GROUP BY ri.drug_name, ri.strength
            ORDER BY total_value DESC
            LIMIT 10
        """, (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        
        top_drugs = cur.fetchall()
        
    except Exception as e:
        app.logger.error(f"Error fetching pharmacy revenue: {e}")
        sales = []
        total_revenue = 0
        avg_transaction = 0
        top_drugs = []
        flash("Error loading pharmacy revenue report", "danger")
    
    cur.close()
    conn.close()
    
    months = [(i, month_name[i]) for i in range(1, 13)]
    years = range(2024, today.year + 1)
    
    return render_template(
        "admin_pharmacy_revenue.html",
        sales=sales,
        total_revenue=total_revenue,
        avg_transaction=avg_transaction,
        top_drugs=top_drugs,
        period=report_type,
        start_date=start_date,
        end_date=end_date,
        selected_day=selected_day,
        selected_month=int(selected_month) if selected_month else today.month,
        selected_year=int(selected_year) if selected_year else today.year,
        months=months,
        years=years,
        today_str=today.strftime("%Y-%m-%d"),
        admin_name=session.get('admin_full_name', 'Admin')
    )
@app.route('/admin/reports/billing-payments')
def admin_billing_payments():
    if 'admin_id' not in session:
        return redirect(url_for('admin_login'))
    
    patient_name = request.args.get("patient_name", "").strip()
    service_type = request.args.get("service_type", "")
    payment_method = request.args.get("payment_method", "")
    status = request.args.get("status", "")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    
    current_filters = {
        'patient_name': patient_name,
        'service_type': service_type,
        'payment_method': payment_method,
        'status': status,
        'start_date': start_date,
        'end_date': end_date
    }
    
    page = request.args.get("page", 1, type=int)
    per_page = 20
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        query = """
            SELECT p.*, bu.username as cashier_name
            FROM payments p
            LEFT JOIN billing_users bu ON p.recorded_by = bu.id
            WHERE 1=1
        """
        count_query = "SELECT COUNT(*) FROM payments WHERE 1=1"
        params = []
        
        if patient_name:
            query += " AND LOWER(p.patient_name) LIKE LOWER(?)"
            count_query += " AND LOWER(patient_name) LIKE LOWER(?)"
            params.append(f"%{patient_name}%")
        
        if service_type:
            query += " AND p.service_type = ?"
            count_query += " AND service_type = ?"
            params.append(service_type)
        
        if payment_method:
            query += " AND p.payment_method = ?"
            count_query += " AND payment_method = ?"
            params.append(payment_method)
        
        if status:
            query += " AND p.status = ?"
            count_query += " AND status = ?"
            params.append(status)
        
        if start_date:
            query += " AND p.payment_date >= ?"
            count_query += " AND payment_date >= ?"
            params.append(start_date)
        
        if end_date:
            query += " AND p.payment_date <= ?"
            count_query += " AND payment_date <= ?"
            params.append(end_date)
        
        cur.execute(count_query, params)
        total_items = cur.fetchone()[0]
        
        query += " ORDER BY p.created_at DESC LIMIT ? OFFSET ?"
        offset = (page - 1) * per_page
        params.extend([per_page, offset])
        
        cur.execute(query, params)
        payments = cur.fetchall()
        
        cur.execute("SELECT DISTINCT service_type FROM payments WHERE service_type IS NOT NULL ORDER BY service_type")
        service_types = [row[0] for row in cur.fetchall()]
        
        total_amount = 0
        formatted_payments = []
        for payment in payments:
            payment_dict = dict(payment)
            payment_dict["amount_paid"] = float(payment_dict["amount_paid"])

            # Parse date/datetime string fields into datetime objects
            for date_field in ("created_at", "payment_date"):
                val = payment_dict.get(date_field)
                if val and isinstance(val, str):
                    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
                        try:
                            payment_dict[date_field] = datetime.strptime(val, fmt)
                            break
                        except ValueError:
                            continue

            formatted_payments.append(payment_dict)
            total_amount += payment_dict["amount_paid"]
        
        stats_query = """
            SELECT 
                COUNT(*) as total_transactions,
                COALESCE(SUM(amount_paid), 0) as total_collected,
                COALESCE(AVG(amount_paid), 0) as avg_payment,
                COUNT(DISTINCT payment_method) as payment_methods_count,
                COUNT(CASE WHEN status = 'Paid' THEN 1 END) as paid_count,
                COUNT(CASE WHEN status = 'Partial' THEN 1 END) as partial_count
            FROM payments
            WHERE 1=1
        """
        stats_params = []
        
        if patient_name:
            stats_query += " AND LOWER(patient_name) LIKE LOWER(?)"
            stats_params.append(f"%{patient_name}%")
        
        if service_type:
            stats_query += " AND service_type = ?"
            stats_params.append(service_type)
        
        if payment_method:
            stats_query += " AND payment_method = ?"
            stats_params.append(payment_method)
        
        if status:
            stats_query += " AND status = ?"
            stats_params.append(status)
        
        if start_date:
            stats_query += " AND payment_date >= ?"
            stats_params.append(start_date)
        
        if end_date:
            stats_query += " AND payment_date <= ?"
            stats_params.append(end_date)
        
        cur.execute(stats_query, stats_params)
        stats_row = cur.fetchone()
        
        if stats_row:
            stats = {
                'total_transactions': stats_row[0],
                'total_collected': float(stats_row[1]) if stats_row[1] else 0.0,
                'avg_payment': float(stats_row[2]) if stats_row[2] else 0.0,
                'payment_methods_count': stats_row[3],
                'paid_count': stats_row[4],
                'partial_count': stats_row[5]
            }
        else:
            stats = {
                'total_transactions': 0,
                'total_collected': 0.0,
                'avg_payment': 0.0,
                'payment_methods_count': 0,
                'paid_count': 0,
                'partial_count': 0
            }
        
    except Exception as e:
        app.logger.error(f"Error fetching billing payments: {e}")
        formatted_payments = []
        service_types = []
        total_items = 0
        total_amount = 0
        stats = {
            'total_transactions': 0,
            'total_collected': 0.0,
            'avg_payment': 0.0,
            'payment_methods_count': 0,
            'paid_count': 0,
            'partial_count': 0
        }
        flash("Error loading billing payments report", "danger")
    
    cur.close()
    conn.close()
    
    total_pages = (total_items + per_page - 1) // per_page
    
    return render_template(
        "admin_billing_payments.html",
        payments=formatted_payments,
        service_types=service_types,
        total_items=total_items,
        total_amount=total_amount,
        stats=stats,
        page=page,
        total_pages=total_pages,
        current_filters=current_filters,
        admin_name=session.get('admin_full_name', 'Admin')
    )

@app.route('/admin/reports/todays-collection')
def admin_todays_collection():
    if 'admin_id' not in session:
        return redirect(url_for('admin_login'))
    
    today = date.today()
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT p.*, bu.username as cashier_name
            FROM payments p
            LEFT JOIN billing_users bu ON p.recorded_by = bu.id
            WHERE DATE(p.payment_date) = ?
            ORDER BY p.created_at DESC
        """, (today.strftime('%Y-%m-%d'),))
        
        today_payments = cur.fetchall()
        
        payment_methods_data = {
            'Cash': {'amount': 0.0, 'count': 0},
            'Card': {'amount': 0.0, 'count': 0},
            'Transfer': {'amount': 0.0, 'count': 0},
            'POS': {'amount': 0.0, 'count': 0},
            'Insurance': {'amount': 0.0, 'count': 0},
            'Other': {'amount': 0.0, 'count': 0}
        }
        
        total_transactions = len(today_payments)
        grand_total = 0.0
        amounts = []
        
        recent_transactions = []
        for payment in today_payments:
            amount_paid = float(payment[7])
            payment_method = payment[9]
            
            grand_total += amount_paid
            amounts.append(amount_paid)
            
            if payment_method in payment_methods_data:
                payment_methods_data[payment_method]['amount'] += amount_paid
                payment_methods_data[payment_method]['count'] += 1
            else:
                payment_methods_data['Other']['amount'] += amount_paid
                payment_methods_data['Other']['count'] += 1
            
            recent_transactions.append({
                'id': payment[0],
                'patient_name': payment[1],
                'service_type': payment[2],
                'amount_paid': amount_paid,
                'payment_method': payment_method,
                'status': payment[10],
                'created_at': payment[13],
                'cashier_name': payment[14]
            })
        
        average_transaction = grand_total / total_transactions if total_transactions > 0 else 0.0
        highest_transaction = max(amounts) if amounts else 0.0
        lowest_transaction = min(amounts) if amounts else 0.0
        
        morning_total = 0.0
        afternoon_total = 0.0
        evening_total = 0.0
        
        for payment in today_payments:
            created_at = payment[13]
            if created_at:
                if isinstance(created_at, str):
                    created_at = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')
                hour = created_at.hour
                amount = float(payment[7])
                
                if 6 <= hour < 12:
                    morning_total += amount
                elif 12 <= hour < 16:
                    afternoon_total += amount
                elif 16 <= hour < 22:
                    evening_total += amount
        
        payment_methods = []
        for method_name, data in payment_methods_data.items():
            if data['count'] > 0:
                percentage = (data['amount'] / grand_total * 100) if grand_total > 0 else 0
                payment_methods.append({
                    'name': method_name,
                    'amount': data['amount'],
                    'count': data['count'],
                    'percentage': round(percentage, 1)
                })
        
        cur.execute("""
            SELECT 
                p.service_type,
                COUNT(*) as transaction_count,
                SUM(p.amount_paid) as total_amount
            FROM payments p
            WHERE DATE(p.payment_date) = ?
            GROUP BY p.service_type
            ORDER BY total_amount DESC
        """, (today.strftime('%Y-%m-%d'),))
        
        service_type_rows = cur.fetchall()
        service_type_data = []
        for row in service_type_rows:
            service_type_data.append({
                'service_type': row[0],
                'transaction_count': row[1],
                'total_amount': float(row[2]) if row[2] else 0.0
            })
        
        daily_target = 500000.00
        progress_percentage = min(100, (grand_total / daily_target * 100)) if daily_target > 0 else 0
        
    except Exception as e:
        app.logger.error(f"Error fetching today's collection: {e}")
        recent_transactions = []
        total_transactions = 0
        grand_total = 0.0
        average_transaction = 0.0
        highest_transaction = 0.0
        lowest_transaction = 0.0
        morning_total = afternoon_total = evening_total = 0.0
        payment_methods = []
        service_type_data = []
        progress_percentage = 0
        flash("Error loading today's collection report", "danger")
    
    cur.close()
    conn.close()
    
    today_date = today.strftime("%A, %B %d, %Y")
    
    return render_template(
        "admin_todays_collection.html",
        today_date=today_date,
        current_time=datetime.now().strftime("%H:%M:%S"),
        grand_total=grand_total,
        cash_total=payment_methods_data['Cash']['amount'],
        card_total=payment_methods_data['Card']['amount'],
        transfer_total=payment_methods_data['Transfer']['amount'],
        pos_total=payment_methods_data['POS']['amount'],
        insurance_total=payment_methods_data['Insurance']['amount'],
        other_total=payment_methods_data['Other']['amount'],
        payment_methods=payment_methods,
        recent_transactions=recent_transactions,
        total_transactions=total_transactions,
        average_transaction=average_transaction,
        highest_transaction=highest_transaction,
        lowest_transaction=lowest_transaction,
        morning_total=morning_total,
        afternoon_total=afternoon_total,
        evening_total=evening_total,
        service_type_data=service_type_data,
        admin_name=session.get('admin_full_name', 'Admin'),
        daily_target=daily_target,
        progress_percentage=progress_percentage
    )

@app.route('/admin/drugs/delete-expired', methods=['POST'])
def admin_delete_expired_drugs():
    if 'admin_id' not in session:
        return redirect(url_for('admin_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT id, name, strength, stock_quantity 
            FROM drugs 
            WHERE expiry_date < date('now') AND stock_quantity > 0
        """)
        
        expired_drugs = cursor.fetchall()
        
        if not expired_drugs:
            flash("No expired drugs found to delete.", "info")
            return redirect(url_for('admin_pharmacy_stock'))
        
        drug_names = [f"{d[1]} {d[2]} (Qty: {d[3]})" for d in expired_drugs]
        
        cursor.execute("DELETE FROM drugs WHERE expiry_date < date('now')")
        
        conn.commit()
        
        log_admin_action(
            session['admin_id'], 
            'DELETE_EXPIRED_DRUGS', 
            f'Removed {len(expired_drugs)} expired drugs: {", ".join(drug_names)}'
        )
        
        flash(f"Successfully removed {len(expired_drugs)} expired drug(s) from database.", "success")
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error deleting expired drugs: {e}")
        flash(f"Error deleting expired drugs: {str(e)}", "danger")
    
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('admin_pharmacy_stock'))

# -------------------- ADMIN EXPORT ROUTES --------------------
@app.route('/admin/reports/export/pharmacy-stock')
def admin_export_pharmacy_stock():
    if 'admin_id' not in session:
        return redirect(url_for('admin_login'))
    
    filter_type = request.args.get("filter", "all")
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT id, name, strength, stock_quantity, unit_price, 
                   expiry_date, low_stock_threshold
            FROM drugs
            ORDER BY expiry_date ASC
        """)
        rows = cur.fetchall()
        
        stock = build_stock_snapshot(rows, date.today())
        stock = apply_stock_filter(stock, filter_type)
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Pharmacy Stock - {filter_type}"
        
        headers = [
            "Drug Name", "Strength", "Quantity", "Unit Price (₦)",
            "Expiry Date", "Days Left", "Status", "Total Value (₦)", "Low Stock Threshold"
        ]
        ws.append(headers)
        
        for c in range(1, len(headers) + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)
        
        fills = {
            "EXPIRED": PatternFill("solid", fgColor="FF9999"),
            "EXPIRING_SOON": PatternFill("solid", fgColor="FFFF99"),
            "LOW": PatternFill("solid", fgColor="ADD8E6")
        }
        
        for item in stock:
            ws.append([
                item["name"], item["strength"], item["quantity"],
                float(item["unit_price"]), item["expiry_date"],
                item["days_left"], item["status"],
                float(item["total_value"]), item["low_stock_threshold"]
            ])
            
            row_idx = ws.max_row
            if item["status"] in fills:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = fills[item["status"]]
            elif item["quantity"] <= item["low_stock_threshold"]:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = fills["LOW"]
        
        ws.append([])
        ws.append(["SUMMARY"])
        ws.append(["Total Items:", len(stock)])
        ws.append(["Total Stock Value:", f"₦{sum(d['total_value'] for d in stock):,.2f}"])
        ws.append(["Expired Items:", sum(1 for d in stock if d["status"] == "EXPIRED")])
        ws.append(["Expiring Soon:", sum(1 for d in stock if d["status"] == "EXPIRING_SOON")])
        ws.append(["Low Stock:", sum(1 for d in stock if d["quantity"] <= d["low_stock_threshold"])])
        
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        log_admin_action(session['admin_id'], 'EXPORT_REPORT', 
                       f'Exported pharmacy stock report (filter: {filter_type})')
        
        return send_file(
            stream,
            as_attachment=True,
            download_name=f"admin_pharmacy_stock_{date.today()}_{filter_type}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        app.logger.error(f"Error exporting pharmacy stock: {e}")
        flash("Error exporting report", "danger")
        return redirect(url_for('admin_pharmacy_stock'))
        
    finally:
        cur.close()
        conn.close()

@app.route('/admin/reports/export/billing-payments')
def admin_export_billing_payments():
    if 'admin_id' not in session:
        return redirect(url_for('admin_login'))
    
    patient_name = request.args.get("patient_name", "").strip()
    service_type = request.args.get("service_type", "")
    payment_method = request.args.get("payment_method", "")
    status = request.args.get("status", "")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        query = """
            SELECT p.*, bu.username as cashier_name
            FROM payments p
            LEFT JOIN billing_users bu ON p.recorded_by = bu.id
            WHERE 1=1
        """
        params = []
        
        if patient_name:
            query += " AND LOWER(p.patient_name) LIKE LOWER(?)"
            params.append(f"%{patient_name}%")
        
        if service_type:
            query += " AND p.service_type = ?"
            params.append(service_type)
        
        if payment_method:
            query += " AND p.payment_method = ?"
            params.append(payment_method)
        
        if status:
            query += " AND p.status = ?"
            params.append(status)
        
        if start_date:
            query += " AND p.payment_date >= ?"
            params.append(start_date)
        
        if end_date:
            query += " AND p.payment_date <= ?"
            params.append(end_date)
        
        query += " ORDER BY p.created_at DESC"
        cur.execute(query, params)
        payments = cur.fetchall()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Billing Payments"
        
        ws.append(["Billing Payments Report"])
        ws.append([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([])
        
        headers = [
            "Receipt No", "Patient Name", "Service Type", 
            "Subtotal (₦)", "Discount (₦)", "Tax (₦)", "Grand Total (₦)",
            "Amount Paid (₦)", "Balance (₦)", "Payment Method",
            "Status", "Payment Date", "Created At", "Cashier"
        ]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        for payment in payments:
            ws.append([
                payment[0], payment[1], payment[2],
                float(payment[3]), float(payment[4]), float(payment[5]),
                float(payment[6]), float(payment[7]), float(payment[8]),
                payment[9], payment[10], payment[11], payment[13], payment[14]
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        total_amount = sum(float(p[7]) for p in payments) if payments else 0
        total_balance = sum(float(p[8]) for p in payments) if payments else 0
        
        ws.append([])
        ws.append(["SUMMARY"])
        ws.append(["Total Transactions:", len(payments)])
        ws.append(["Total Amount Collected:", total_amount])
        ws.append(["Total Outstanding Balance:", total_balance])
        
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        log_admin_action(session['admin_id'], 'EXPORT_REPORT', 
                       f'Exported billing payments report')
        
        return send_file(
            stream,
            as_attachment=True,
            download_name=f"admin_billing_payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        app.logger.error(f"Error exporting billing payments: {e}")
        flash("Error exporting report", "danger")
        return redirect(url_for('admin_billing_payments'))
        
    finally:
        cur.close()
        conn.close()

@app.route('/admin/reports/export/todays-collection')
def admin_export_todays_collection():
    if 'admin_id' not in session:
        return redirect(url_for('admin_login'))
    
    today = date.today()
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT p.*, bu.username as cashier_name
            FROM payments p
            LEFT JOIN billing_users bu ON p.recorded_by = bu.id
            WHERE DATE(p.payment_date) = ?
            ORDER BY p.created_at DESC
        """, (today.strftime('%Y-%m-%d'),))
        
        today_payments = cur.fetchall()
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Today's Collection - {today}"
        
        ws.append([f"Today's Collection Report - {today.strftime('%B %d, %Y')}"])
        ws.append([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([])
        
        ws.append(["SUMMARY"])
        ws.append([])
        
        grand_total = sum(float(p[7]) for p in today_payments) if today_payments else 0
        total_transactions = len(today_payments)
        
        ws.append(["Date:", today.strftime('%A, %B %d, %Y')])
        ws.append(["Total Transactions:", total_transactions])
        ws.append(["Grand Total:", grand_total])
        ws.append([])
        
        payment_methods_data = {
            'Cash': {'amount': 0, 'count': 0},
            'Card': {'amount': 0, 'count': 0},
            'Transfer': {'amount': 0, 'count': 0},
            'POS': {'amount': 0, 'count': 0},
            'Insurance': {'amount': 0, 'count': 0},
            'Other': {'amount': 0, 'count': 0}
        }
        
        for payment in today_payments:
            amount_paid = float(payment[7])
            payment_method = payment[9]
            
            if payment_method in payment_methods_data:
                payment_methods_data[payment_method]['amount'] += amount_paid
                payment_methods_data[payment_method]['count'] += 1
            else:
                payment_methods_data['Other']['amount'] += amount_paid
                payment_methods_data['Other']['count'] += 1
        
        ws.append(["Payment Method Breakdown"])
        ws.append(["Method", "Count", "Amount", "Percentage"])
        
        for method_name, data in payment_methods_data.items():
            if data['count'] > 0:
                percentage = (data['amount'] / grand_total * 100) if grand_total > 0 else 0
                ws.append([
                    method_name,
                    data['count'],
                    data['amount'],
                    f"{percentage:.1f}%"
                ])
        
        ws.append([])
        ws.append([])
        ws.append(["DETAILED TRANSACTIONS"])
        ws.append([])
        
        headers = [
            "Receipt No", "Patient Name", "Service Type", 
            "Subtotal", "Discount", "Tax", "Grand Total",
            "Amount Paid", "Balance", "Payment Method",
            "Status", "Payment Date", "Time", "Cashier"
        ]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), 
                       right=Side(style='thin'), 
                       top=Side(style='thin'), 
                       bottom=Side(style='thin'))
        
        for cell in ws[ws.max_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        for payment in today_payments:
            created_at = payment[13]
            time_str = created_at.strftime('%H:%M:%S') if hasattr(created_at, 'strftime') else str(created_at)
            
            ws.append([
                payment[0], payment[1], payment[2],
                float(payment[3]), float(payment[4]), float(payment[5]),
                float(payment[6]), float(payment[7]), float(payment[8]),
                payment[9], payment[10], payment[11], time_str, payment[14]
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        log_admin_action(session['admin_id'], 'EXPORT_REPORT', 
                       f'Exported today\'s collection report')
        
        return send_file(
            stream,
            as_attachment=True,
            download_name=f"todays_collection_{today.strftime('%Y%m%d')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        app.logger.error(f"Error exporting today's collection: {e}")
        flash("Error exporting report", "danger")
        return redirect(url_for('admin_todays_collection'))
    
    finally:
        cur.close()
        conn.close()

# -------------------- MAIN --------------------
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    
    create_tables()
    create_default_users()
    create_hr_tables()
    create_default_admin()
    create_nkiru_user()    # Create Nkiru (Cashier1)
    create_christy_user()
    add_missing_columns()
    sync_existing_users()
    
    debug_mode = os.environ.get("FLASK_DEBUG", "False").lower() == "true"
    app.run(host="0.0.0.0", port=port, debug=debug_mode)